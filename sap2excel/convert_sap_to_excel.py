#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SAP CSV to Excel Converter (Strict CSV Order & Style Cloning)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os

# ================= CONFIGURATION =================
EXCEL_HEADER_ROW = 5   # บรรทัดหัวตารางใน Template
DATA_START_ROW = 6     # บรรทัดเริ่มเขียนข้อมูล

CSV_MAPPING = [
    {
        'file': '001_ต้นทุน_BU_1025 - 10-11-68.csv',
        'encoding': 'cp874',
        'sheet': 'ต้นทุน_กลุ่มธุรกิจ'
    },
    {
        'file': '001_ต้นทุน_Product_1025 - 10-11-68.csv',
        'encoding': 'cp874',
        'sheet': 'ต้นทุน_กลุ่มบริการ'
    },
    {
        'file': '001_ต้นทุน_Product_Group_1025 - 10-11-68.csv',
        'encoding': 'cp874',
        'sheet': 'ต้นทุน_บริการ'
    },
    {
        'file': '002_บัญชี_BU_1025 - 10-11-68.csv',
        'encoding': 'cp874',
        'sheet': 'หมวดบัญชี_กลุ่มธุรกิจ'
    },
    {
        'file': '002_บัญชี_Product_1025 - 10-11-68.csv',
        'encoding': 'cp874',
        'sheet': 'หมวดบัญชี_กลุ่มบริการ'
    },
    {
        'file': '002_บัญชี_Product_Grop_1025 - 10-11-68.csv',
        'encoding': 'cp874',
        'sheet': 'หมวดบัญชี_บริการ'
    }
]
# =================================================

def clean_sap_value(value):
    """แปลงค่าตัวเลขจาก SAP (1,234.00-) เป็น float (-1234.00)"""
    if pd.isna(value) or str(value).strip() == '':
        return None
    
    s_val = str(value).strip().replace(',', '') # เอา comma ออก
    
    # จัดการเครื่องหมายลบข้างหลัง
    if s_val.endswith('-'):
        try:
            return -float(s_val[:-1])
        except:
            return value
    
    # แปลงตัวเลขปกติ
    try:
        return float(s_val)
    except:
        return value

def find_csv_header_row(file_path, encoding):
    """หาบรรทัด Header ของ CSV"""
    try:
        with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
            for i, line in enumerate(f):
                # SAP Report มักขึ้นต้นด้วย 'รายละเอียด' หรือ 'Account'
                if 'รายละเอียด' in line or 'Account' in line or 'Description' in line:
                    return i
    except:
        pass
    return 0 # Default

def read_csv_file(file_path, encoding):
    print(f"  📄 Reading: {os.path.basename(file_path)}")
    
    header_row = find_csv_header_row(file_path, encoding)
    print(f"     > Found header at row: {header_row + 1}")
    
    try:
        # ลองอ่านด้วย Tab (\t)
        df = pd.read_csv(file_path, sep='\t', encoding=encoding, header=header_row, on_bad_lines='skip')
        if len(df.columns) <= 1:
             # ถ้าไม่เวิร์ค ลอง Comma
             df = pd.read_csv(file_path, sep=',', encoding=encoding, header=header_row, on_bad_lines='skip')
    except:
        df = pd.read_csv(file_path, sep=',', encoding=encoding, header=header_row, on_bad_lines='skip')

    # ลบคอลัมน์ขยะ
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    df.columns = df.columns.str.strip() # ลบช่องว่างชื่อหัวตาราง
    
    # Clean Data (แปลงตัวเลข)
    for col in df.columns:
        if col != df.columns[0]: # เว้นคอลัมน์แรก (ชื่อรายการ)
            df[col] = df[col].apply(clean_sap_value)
            
    print(f"     > Data Loaded: {len(df)} rows, {len(df.columns)} columns")
    return df

def apply_style(cell, template_cell, is_number=True):
    """ฟังก์ชันก๊อปปี้ Style จากเซลล์ต้นแบบ"""
    if template_cell and template_cell.has_style:
        cell.font = copy(template_cell.font)
        cell.border = copy(template_cell.border)
        cell.fill = copy(template_cell.fill)
        cell.alignment = copy(template_cell.alignment)
        
        # ถ้าเป็นตัวเลข ให้ใช้ Format บัญชีเสมอ (ไม่ก๊อปปี้ format เดิม)
        if is_number:
            cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
        else:
            cell.number_format = copy(template_cell.number_format)

def write_and_format(ws, df):
    """เขียนข้อมูลเรียงตาม CSV และ Clone Style"""
    print(f"     > Writing to sheet: {ws.title}")
    
    # 1. จำ Style ต้นแบบ (จากแถวข้อมูลแถวแรกของ Template)
    # เราจะใช้คอลัมน์ 1 เป็นต้นแบบ Text และคอลัมน์ 2 เป็นต้นแบบ Number
    style_template_text = ws.cell(row=DATA_START_ROW, column=1)
    style_template_num = ws.cell(row=DATA_START_ROW, column=2)
    
    # จำ Style หัวตารางด้วย (จากคอลัมน์ 2 แถว 5)
    style_header_num = ws.cell(row=EXCEL_HEADER_ROW, column=2)

    # 2. ลบข้อมูลเก่าทิ้ง (Clear Data) แต่เก็บ Header ไว้
    # ลบตั้งแต่แถวข้อมูลลงไปจนสุด และลบคอลัมน์ขวาทิ้งทั้งหมดเพื่อเขียนใหม่
    ws.delete_rows(DATA_START_ROW, ws.max_row)
    
    # 3. เขียน Header ใหม่ (ตาม CSV)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=EXCEL_HEADER_ROW, column=col_idx)
        cell.value = col_name
        
        # จัด Format Header (ถ้าคอลัมน์เกิน Template เดิม ให้ก๊อปจากต้นแบบ)
        if col_idx > 1:
            apply_style(cell, style_header_num, is_number=False)
            
        # ปรับความกว้างคอลัมน์ (ถ้าเป็นคอลัมน์ใหม่)
        col_letter = get_column_letter(col_idx)
        if not ws.column_dimensions[col_letter].width:
             ws.column_dimensions[col_letter].width = 15 # default width

    # 4. เขียนข้อมูล (Data)
    for r_idx, row in enumerate(df.values, start=DATA_START_ROW):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
            
            # 5. ใส่ Style
            if c_idx == 1:
                # คอลัมน์แรก (Text/Description)
                apply_style(cell, style_template_text, is_number=False)
            else:
                # คอลัมน์อื่นๆ (ตัวเลข)
                apply_style(cell, style_template_num, is_number=True)

    print(f"     > Success! Wrote {len(df)} rows.")

def main(csv_dir, template_path, output_path):
    print("="*60)
    print("🚀 STARTING CONVERSION (STRICT CSV ORDER)")
    print("="*60)

    print(f"📂 Loading Template: {os.path.basename(template_path)}")
    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception as e:
        print(f"❌ Error loading template: {e}")
        return

    import glob
    for mapping in CSV_MAPPING:
        print("-" * 40)
        
        # Use glob to find the file that matches the pattern
        file_pattern = mapping['file'].split('_1025')[0] + '*.csv'
        found_files = glob.glob(os.path.join(csv_dir, file_pattern))
        
        if not found_files:
            print(f"⚠️  Skipping: Pattern '{file_pattern}' not found.")
            continue
            
        csv_path = found_files[0] # Use the first file found
        sheet_name = mapping['sheet']

        if not os.path.exists(csv_path):
            print(f"⚠️  Skipping: {os.path.basename(csv_path)} (Not Found)")
            continue

        # อ่าน CSV
        try:
            df = read_csv_file(csv_path, mapping['encoding'])
        except Exception as e:
            print(f"❌ Error reading CSV: {e}")
            continue

        # เขียนลง Excel
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                write_and_format(ws, df)
            except Exception as e:
                print(f"❌ Error writing: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"⚠️  Sheet '{sheet_name}' not found. Creating new.")
            ws = wb.create_sheet(sheet_name)
            # เขียนแบบไม่มี Style อ้างอิง (เพราะไม่มี Template Sheet นี้)
            from openpyxl.utils.dataframe import dataframe_to_rows
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

    print("="*60)
    print(f"💾 Saving to: {output_path}")
    wb.save(output_path)
    print("✅ DONE.")

if __name__ == "__main__":
    # --- Paths are now relative to the script's location ---
    BASE_DIR = "." 
    
    TEMPLATE_FILE = os.path.join(BASE_DIR, "Report_NT BU_สะสม 202510_T.xlsx")
    OUTPUT_FILE = os.path.join(BASE_DIR, "Report_NT_BU_Final_Ordered.xlsx")

    if os.path.exists(TEMPLATE_FILE):
        main(BASE_DIR, TEMPLATE_FILE, OUTPUT_FILE)
    else:
        print(f"❌ Template file not found: {TEMPLATE_FILE}")