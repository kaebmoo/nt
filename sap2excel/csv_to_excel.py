#!/usr/bin/env python3
"""
CSV to Excel Converter with Formatting
สคริปต์นี้อ่านไฟล์ CSV จาก SAP และแปลงเป็น Excel ที่มีการจัดรูปแบบสวยงาม
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import re


class CSVToExcelConverter:
    """แปลงไฟล์ CSV เป็น Excel พร้อมจัดรูปแบบ"""

    def __init__(self, base_path="."):
        self.base_path = base_path

        # กำหนดการ mapping ระหว่าง CSV และ Sheet name
        self.file_mapping = {
            "001_ต้นทุน_BU": "ต้นทุน_กลุ่มธุรกิจ",
            "001_ต้นทุน_Product_Group": "ต้นทุน_กลุ่มบริการ",
            "001_ต้นทุน_Product": "ต้นทุน_บริการ",
            "002_บัญชี_BU": "หมวดบัญชี_กลุ่มธุรกิจ",
            "002_บัญชี_Product_Grop": "หมวดบัญชี_กลุ่มบริการ",  # ใช้ชื่อเดิมที่มี typo
            "002_บัญชี_Product": "หมวดบัญชี_บริการ",
        }

        # Styles
        self.font_header = Font(name='TH Sarabun New', size=18, bold=True)
        self.font_data = Font(name='TH Sarabun New', size=16)
        self.fill_header = PatternFill(start_color='FFF4DEDC', end_color='FFF4DEDC', fill_type='solid')
        self.alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.alignment_vcenter = Alignment(vertical='center')
        self.alignment_right = Alignment(horizontal='right', vertical='center')

        # สีสำหรับ BU Headers (Column) - กำหนดลำดับความสำคัญในการจับคู่
        self.bu_colors = {
            'HARD INFRASTRUCTURE': 'FFE2EFDA',
            'INTERNATIONAL': 'FFDDEBF7',
            'MOBILE': 'FFDBD3E5',
            'FIXED LINE & BROADBAND': 'FFFCE4D6',
            'DIGITAL': 'FFD9E1F2',
            'ICT SOLUTION': 'FFC6E0B4',
            'อื่นไม่ใช่โทรคมนาคม': 'FFBDD7EE',
            'รายได้อื่น/ค่าใช้จ่ายอื่น': 'FFEAC1C0',
        }

        # สีสำหรับ Row Headers (รายละเอียด)
        self.fill_description = PatternFill(start_color='FFF4DEDC', end_color='FFF4DEDC', fill_type='solid')
        self.fill_main_row = PatternFill(start_color='FFF8CBAD', end_color='FFF8CBAD', fill_type='solid')

        # Borders
        self.border_thin = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        self.border_top_only = Border(top=Side(style='thin'))
        self.border_bottom_only = Border(bottom=Side(style='thin'))

    def find_csv_files(self, period):
        """ค้นหาไฟล์ CSV ตาม period (เช่น 1025)"""
        csv_files = {}

        for file_prefix, sheet_name in self.file_mapping.items():
            # ค้นหาไฟล์ที่ตรงกับ pattern
            pattern = f"{file_prefix}_{period}"

            for filename in os.listdir(self.base_path):
                if filename.startswith(pattern) and filename.endswith('.csv'):
                    csv_files[sheet_name] = os.path.join(self.base_path, filename)
                    break

        return csv_files

    def read_csv_data(self, csv_path):
        """อ่านไฟล์ CSV และแยกส่วนต่างๆ"""
        # อ่านไฟล์ CSV ด้วย encoding cp874 (Thai Windows) และ tab delimiter
        df = pd.read_csv(csv_path, encoding='cp874', sep='\t')

        # ดึงข้อมูลจาก header
        report_title_raw = df.iloc[0, 0] if len(df) > 0 else "รายงานผลการดำเนินงาน"
        # แยกชื่อรายงาน (มักจะมี tab characters ต่อท้าย)
        if isinstance(report_title_raw, str):
            report_title = report_title_raw.split('\t')[0].strip()
        else:
            report_title = "รายงานผลการดำเนินงาน"

        date_printed = df.iloc[1, 0] if len(df) > 1 else ""

        # แยกวันที่ออกมา
        date_match = re.search(r'\d{2}-\d{2}-\d{4}\s+\d{2}:\d{2}', str(date_printed))
        if date_match:
            date_str = date_match.group()
        else:
            date_str = datetime.now().strftime("%d-%m-%Y %H:%M")

        # โครงสร้าง CSV:
        # Row 0 = ว่าง
        # Row 1 = วัน เวลาที่พิมพ์
        # Row 2 = ว่าง
        # Row 3 = Business Unit headers (รายละเอียด, รวมทั้งสิ้น, BU ต่างๆ)
        # Row 4 = Sub headers (รวม กลุ่มธุรกิจ... หรือ Product Group)
        # Row 5 = Product Code (สำหรับ Product sheets)
        # Row 6 = Product Name (สำหรับ Product sheets)
        # Row 7+ = Data (ข้อมูลจริง)

        # ตรวจสอบว่าเป็น Product sheet หรือไม่
        has_product_headers = False
        if len(df) > 5:
            # ถ้า row 5 มีข้อมูลที่ไม่ใช่ NaN แสดงว่าเป็น Product sheet
            row5_values = df.iloc[5].tolist()
            if any(pd.notna(v) and str(v).strip() != '' for v in row5_values):
                has_product_headers = True

        # แก้ไขตัวเลขที่มีเครื่องหมายลบต่อท้าย เช่น "419523515.74-" เป็น "-419523515.74"
        data_df = df.iloc[7:].reset_index(drop=True) if len(df) > 7 else pd.DataFrame()

        # แปลงตัวเลขที่มี - ต่อท้ายให้เป็นเลขลบปกติ
        for col in data_df.columns:
            if col != 0:  # ไม่ต้องแปลง column แรก (รายละเอียด)
                data_df[col] = data_df[col].apply(self._fix_negative_number)

        result = {
            'report_title': report_title,
            'date_printed': date_str,
            'bu_headers': df.iloc[3].tolist() if len(df) > 3 else [],
            'sub_headers': df.iloc[4].tolist() if len(df) > 4 else [],
            'has_product_headers': has_product_headers,
            'data_df': data_df,
            'full_df': df
        }

        # เพิ่ม Product headers ถ้ามี
        if has_product_headers:
            result['product_code_headers'] = df.iloc[5].tolist() if len(df) > 5 else []
            result['product_name_headers'] = df.iloc[6].tolist() if len(df) > 6 else []

        return result

    def _fix_negative_number(self, value):
        """แปลงเลขที่มีเครื่องหมายลบต่อท้าย เช่น '419523515.74-' เป็น -419523515.74
        และแปลง string ตัวเลขเป็น float
        """
        if isinstance(value, str):
            value = value.strip()

            # ถ้าเป็น string ว่างหรือไม่ใช่ตัวเลข ให้คืนค่าเดิม
            if value == '' or value == 'nan':
                return value

            # ตรวจสอบว่ามีเครื่องหมาย - ต่อท้ายหรือไม่
            if value.endswith('-'):
                # ลบเครื่องหมาย - และแปลงเป็นตัวเลข
                try:
                    num = float(value[:-1].replace(',', ''))
                    return -num
                except:
                    return value
            else:
                # พยายามแปลงเป็นตัวเลข
                try:
                    num = float(value.replace(',', ''))
                    return num
                except:
                    return value

        return value

    def get_period_text(self, period):
        """แปลง period (เช่น 1025) เป็น text (เช่น 10 เดือน สิ้นสุดวันที่ 31 ตุลาคม 2568)

        Format: MMYY
        - MM = เดือน (01-12)
        - YY = ปี ค.ศ. (เช่น 25 = 2025)
        """
        if not period:
            return "สำหรับงวด..."

        month_dict = {
            '01': ('มกราคม', 31), '02': ('กุมภาพันธ์', 28), '03': ('มีนาคม', 31),
            '04': ('เมษายน', 30), '05': ('พฤษภาคม', 31), '06': ('มิถุนายน', 30),
            '07': ('กรกฎาคม', 31), '08': ('สิงหาคม', 31), '09': ('กันยายน', 30),
            '10': ('ตุลาคม', 31), '11': ('พฤศจิกายน', 30), '12': ('ธันวาคม', 31)
        }

        # Parse period (format: MMYY)
        if len(period) >= 4:
            month_num = period[:2]
            year_suffix = period[2:4]

            # แปลงปี ค.ศ. เป็น พ.ศ. (เช่น 25 = 2025 ค.ศ. = 2568 พ.ศ.)
            year_int = int(year_suffix)
            year_ce = 2000 + year_int  # ค.ศ.
            year_be = year_ce + 543    # พ.ศ. = ค.ศ. + 543

            if month_num in month_dict:
                month_name, last_day = month_dict[month_num]
                month_int = int(month_num)

                return f"สำหรับงวด {month_int} เดือน สิ้นสุดวันที่ {last_day} {month_name} {year_be}"

        return "สำหรับงวด..."

    def create_formatted_sheet(self, ws, sheet_data, period):
        """สร้างและจัดรูปแบบ sheet"""

        # บรรทัดที่ 1: ว่าง
        # บรรทัดที่ 2: ชื่อบริษัท
        ws['B2'] = 'บริษัท โทรคมนาคมแห่งชาติ จำกัด (มหาชน)'
        ws['B2'].font = self.font_header
        ws['B2'].alignment = self.alignment_vcenter
        # Merge cells สำหรับ header (แนวนอน)
        ws.merge_cells('B2:E2')

        # บรรทัดที่ 3: ชื่อรายงาน
        report_title = sheet_data['report_title']
        ws['B3'] = report_title
        ws['B3'].font = self.font_header
        ws['B3'].alignment = self.alignment_vcenter
        # Merge cells สำหรับ report title (แนวนอน)
        ws.merge_cells('B3:E3')

        # บรรทัดที่ 4: งวดเวลา
        period_text = self.get_period_text(period)
        ws['B4'] = period_text
        ws['B4'].font = self.font_header
        ws['B4'].alignment = self.alignment_vcenter
        # Merge cells สำหรับ period (แนวนอน)
        ws.merge_cells('B4:E4')

        # บรรทัดที่ 5: ว่าง

        # บรรทัดที่ 6: Business Unit Headers
        bu_color_map = {}  # เก็บ mapping ระหว่าง column กับสี BU
        bu_headers = sheet_data['bu_headers']
        if bu_headers:
            # ค้นหา BU headers ที่ไม่ซ้ำและตำแหน่งที่ต้อง merge
            bu_merge_ranges = []
            current_bu = None
            start_col = None

            for col_idx, header in enumerate(bu_headers, start=2):
                header_str = str(header).strip() if pd.notna(header) else ''

                # ถ้าเจอ header ใหม่ที่ไม่ว่าง
                if header_str != '' and header_str != 'nan':
                    # ถ้ามี BU ก่อนหน้า ให้บันทึก merge range
                    if current_bu is not None and start_col is not None:
                        bu_merge_ranges.append((start_col, col_idx - 1, current_bu))

                    # เริ่ม BU ใหม่
                    current_bu = header_str
                    start_col = col_idx
                else:
                    # ถ้าเป็นคอลัมน์ว่าง และมี current_bu อยู่ แสดงว่ายังอยู่ในช่วงเดียวกัน
                    pass

            # บันทึก BU สุดท้าย
            if current_bu is not None and start_col is not None:
                bu_merge_ranges.append((start_col, len(bu_headers) + 1, current_bu))

            # วาง BU headers และ merge cells พร้อมบันทึก mapping สี
            for start_col, end_col, bu_name in bu_merge_ranges:
                # หาสีสำหรับ BU นี้
                bu_color = self._get_bu_color(bu_name)
                bu_fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type='solid')

                # บันทึก mapping ของสีสำหรับ columns ในช่วงนี้
                for col in range(start_col, end_col + 1):
                    bu_color_map[col] = bu_color

                # วางค่าในเซลล์แรก
                cell = ws.cell(row=6, column=start_col, value=bu_name)
                cell.font = self.font_header
                cell.fill = bu_fill
                cell.alignment = self.alignment_center
                cell.border = self.border_thin

                # Merge cells ถ้ามีมากกว่า 1 คอลัมน์
                if end_col > start_col:
                    start_letter = get_column_letter(start_col)
                    end_letter = get_column_letter(end_col)
                    ws.merge_cells(f'{start_letter}6:{end_letter}6')

                    # ใส่ border และสีให้กับเซลล์ที่ merge ด้วย
                    for col in range(start_col, end_col + 1):
                        ws.cell(row=6, column=col).border = self.border_thin
                        ws.cell(row=6, column=col).fill = bu_fill

        # บรรทัดที่ 7: ว่าง

        # บรรทัดที่ 8: Sub Headers (Product Group - ใช้สีตาม BU และ merge cells)
        sub_headers = sheet_data['sub_headers']
        if sub_headers:
            # หา Sub header ranges ที่ต้อง merge (เหมือนกับ BU headers)
            sub_merge_ranges = []
            current_sub = None
            start_col = None

            for col_idx, header in enumerate(sub_headers, start=2):
                header_str = str(header).strip() if pd.notna(header) else ''

                # ถ้าเจอ header ใหม่ที่ไม่ว่าง
                if header_str != '' and header_str != 'nan':
                    # ถ้ามี sub header ก่อนหน้า ให้บันทึก merge range
                    if current_sub is not None and start_col is not None:
                        sub_merge_ranges.append((start_col, col_idx - 1, current_sub))

                    # เริ่ม sub header ใหม่
                    current_sub = header_str
                    start_col = col_idx
                else:
                    # ถ้าเป็นคอลัมน์ว่าง และมี current_sub อยู่ แสดงว่ายังอยู่ในช่วงเดียวกัน
                    pass

            # บันทึก sub header สุดท้าย
            if current_sub is not None and start_col is not None:
                sub_merge_ranges.append((start_col, len(sub_headers) + 1, current_sub))

            # วาง Sub headers และ merge cells
            for start_col, end_col, sub_name in sub_merge_ranges:
                # หาสีจาก BU color map
                sub_color = bu_color_map.get(start_col, 'FFF4DEDC')
                sub_fill = PatternFill(start_color=sub_color, end_color=sub_color, fill_type='solid')

                # วางค่าในเซลล์แรก
                cell = ws.cell(row=8, column=start_col, value=sub_name)
                cell.font = Font(name='TH Sarabun New', size=14, bold=True)
                cell.fill = sub_fill
                cell.alignment = self.alignment_center
                cell.border = self.border_thin

                # Merge cells ถ้ามีมากกว่า 1 คอลัมน์
                if end_col > start_col:
                    start_letter = get_column_letter(start_col)
                    end_letter = get_column_letter(end_col)
                    ws.merge_cells(f'{start_letter}8:{end_letter}8')

                    # ใส่ border และสีให้กับเซลล์ที่ merge ด้วย
                    for col in range(start_col, end_col + 1):
                        ws.cell(row=8, column=col).border = self.border_thin
                        ws.cell(row=8, column=col).fill = sub_fill

        # บรรทัดที่ 9: Product Code Headers (ถ้ามี)
        # บรรทัดที่ 10: Product Name Headers (ถ้ามี)
        has_product_headers = sheet_data.get('has_product_headers', False)
        if has_product_headers:
            # Row 9: Product Code
            product_code_headers = sheet_data.get('product_code_headers', [])
            if product_code_headers:
                for col_idx, header in enumerate(product_code_headers, start=2):
                    cell = ws.cell(row=9, column=col_idx, value=header if pd.notna(header) and str(header).strip() != '' else '')
                    cell.font = Font(name='TH Sarabun New', size=14, bold=True)
                    cell.alignment = self.alignment_center
                    cell.border = self.border_thin
                    # ใส่สีตาม BU (ทุกคอลัมน์ รวม Total และ cells ว่าง)
                    if col_idx in bu_color_map:
                        bu_color = bu_color_map[col_idx]
                        cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type='solid')

            # Row 10: Product Name
            product_name_headers = sheet_data.get('product_name_headers', [])
            if product_name_headers:
                for col_idx, header in enumerate(product_name_headers, start=2):
                    cell = ws.cell(row=10, column=col_idx, value=header if pd.notna(header) and str(header).strip() != '' else '')
                    cell.font = Font(name='TH Sarabun New', size=14, bold=True)
                    cell.alignment = self.alignment_center
                    cell.border = self.border_thin
                    # ใส่สีตาม BU (ทุกคอลัมน์ รวม Total และ cells ว่าง)
                    if col_idx in bu_color_map:
                        bu_color = bu_color_map[col_idx]
                        cell.fill = PatternFill(start_color=bu_color, end_color=bu_color, fill_type='solid')

        # กำหนด start_row ตามประเภท sheet
        if has_product_headers:
            start_row = 11  # ถ้ามี Product headers ให้เริ่มที่แถว 11
        else:
            start_row = 10  # ถ้าไม่มีให้เริ่มที่แถว 10

        # Data
        data_df = sheet_data['data_df']  # ข้อมูลที่เริ่มจาก row 7 ของ CSV

        for row_idx, row_data in data_df.iterrows():
            excel_row = start_row + row_idx

            # ตรวจสอบประเภทของแถว
            first_col_value = row_data.iloc[0] if len(row_data) > 0 else ''
            first_col_str = str(first_col_value).strip() if pd.notna(first_col_value) else ''

            # แถวหลัก (01., 02., etc.)
            is_main_row = (len(first_col_str) >= 3 and
                          first_col_str[:2].isdigit() and
                          first_col_str[2] == '.')

            # แถวที่ขึ้นต้นด้วย # (ไม่ต้องระบายสี)
            is_hash_row = first_col_str.startswith('#')

            for col_idx, value in enumerate(row_data, start=2):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)

                # จัดรูปแบบตามประเภทข้อมูล
                if col_idx == 2:  # Column B = รายละเอียด
                    cell.font = Font(name='TH Sarabun New', size=16, bold=True if is_hash_row else False)
                    cell.alignment = self.alignment_vcenter
                    cell.border = self.border_thin
                    # ใส่สีพื้นหลัง (ยกเว้นแถวที่ขึ้นต้นด้วย #)
                    if not is_hash_row:
                        if is_main_row:
                            cell.fill = self.fill_main_row  # สีส้มอ่อนสำหรับรายการหลัก
                        else:
                            cell.fill = self.fill_description  # สีชมพูอ่อนสำหรับรายการอื่น
                else:
                    # ตัวเลข
                    if isinstance(value, (int, float)):
                        cell.font = Font(name='TH Sarabun New', size=16)
                        cell.alignment = self.alignment_right

                        # ถ้าค่าเป็น 0 หรือใกล้ 0 ให้แสดงเป็นช่องว่าง
                        if abs(value) < 0.01:
                            cell.value = ''
                        else:
                            # ใช้รูปแบบบัญชี: เลขลบแสดงเป็น (xxx.xx) สีแดง
                            if value < 0:
                                cell.number_format = '#,##0.00_);[Red](#,##0.00)'
                                cell.font = Font(name='TH Sarabun New', size=16, color='FF0000')
                            else:
                                cell.number_format = '#,##0.00_);[Red](#,##0.00)'

                        cell.border = self.border_thin
                        # ใส่สีพื้นหลังสำหรับรายการหลัก (ยกเว้นแถวที่ขึ้นต้นด้วย #)
                        if is_main_row and not is_hash_row:
                            cell.fill = self.fill_main_row
                    else:
                        cell.font = Font(name='TH Sarabun New', size=16)
                        cell.alignment = self.alignment_center
                        cell.border = self.border_thin
                        # ใส่สีพื้นหลังสำหรับรายการหลัก (ยกเว้นแถวที่ขึ้นต้นด้วย #)
                        if is_main_row and not is_hash_row:
                            cell.fill = self.fill_main_row

        # Merge cells แนวตั้งสำหรับ column B (รายละเอียด) - merge cells ที่มีค่าเหมือนกันติดกัน
        # แต่ข้ามกรณีที่เป็น header ที่ขึ้นต้นด้วย # หรือ 01., 02., etc.
        self._merge_description_column(ws, start_row, len(data_df))

        # ปรับความกว้างของคอลัมน์
        ws.column_dimensions['B'].width = 65
        for col_idx in range(3, 50):  # Columns C onwards
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15

        # ปรับความสูงของแถว
        ws.row_dimensions[2].height = 21.65
        ws.row_dimensions[3].height = 21.65
        ws.row_dimensions[4].height = 21.65
        ws.row_dimensions[6].height = 14.15
        ws.row_dimensions[9].height = 54.0

    def _get_bu_color(self, bu_name):
        """หาสีสำหรับ BU จากชื่อ

        Args:
            bu_name: ชื่อ BU

        Returns:
            สี HEX code
        """
        # ตรวจสอบว่า bu_name มีคำสำคัญที่ตรงกับ key ใน bu_colors หรือไม่
        for bu_key, color in self.bu_colors.items():
            if bu_key in bu_name:
                return color

        # ถ้าไม่เจอ ให้ใช้สีเริ่มต้น
        return 'FFF4DEDC'

    def _merge_description_column(self, ws, start_row, num_rows):
        """Merge cells แนวตั้งใน column B สำหรับค่าที่เหมือนกันติดกัน

        Args:
            ws: worksheet
            start_row: แถวเริ่มต้นของข้อมูล
            num_rows: จำนวนแถวข้อมูล
        """
        if num_rows == 0:
            return

        merge_ranges = []
        current_value = None
        merge_start = None

        for row_idx in range(num_rows):
            excel_row = start_row + row_idx
            cell_value = ws.cell(row=excel_row, column=2).value  # Column B

            # แปลงค่าเป็น string และตัดช่องว่าง
            value_str = str(cell_value).strip() if cell_value is not None else ''

            # ไม่ merge กรณีที่เป็น:
            # 1. Header ที่ขึ้นต้นด้วยตัวเลข (01., 02., etc.)
            # 2. Header ที่ขึ้นต้นด้วย #
            # 3. ค่าว่าง
            is_header = (value_str.startswith('#') or
                        (len(value_str) >= 3 and value_str[:2].isdigit() and value_str[2] == '.'))

            if is_header or value_str == '':
                # บันทึก merge range ก่อนหน้า (ถ้ามี)
                if merge_start is not None and excel_row > merge_start + 1:
                    merge_ranges.append((merge_start, excel_row - 1))

                # Reset
                current_value = None
                merge_start = None
            else:
                # ถ้าค่าเหมือนกับค่าก่อนหน้า
                if value_str == current_value:
                    # ยังอยู่ในช่วงเดียวกัน
                    pass
                else:
                    # บันทึก merge range ก่อนหน้า (ถ้ามี)
                    if merge_start is not None and excel_row > merge_start + 1:
                        merge_ranges.append((merge_start, excel_row - 1))

                    # เริ่มช่วงใหม่
                    current_value = value_str
                    merge_start = excel_row

        # บันทึก merge range สุดท้าย
        if merge_start is not None and start_row + num_rows > merge_start + 1:
            merge_ranges.append((merge_start, start_row + num_rows - 1))

        # ทำการ merge cells
        for start_r, end_r in merge_ranges:
            if end_r > start_r:  # มีอย่างน้อย 2 แถว
                ws.merge_cells(f'B{start_r}:B{end_r}')
                # ตั้งค่า alignment ให้ center vertically
                ws.cell(row=start_r, column=2).alignment = self.alignment_vcenter

    def convert(self, period, output_filename=None):
        """แปลงไฟล์ CSV เป็น Excel

        Args:
            period: งวดเวลา (เช่น "1025" สำหรับ ตุลาคม 2568)
            output_filename: ชื่อไฟล์ output (ถ้าไม่ระบุจะสร้างอัตโนมัติ)
        """
        # ค้นหาไฟล์ CSV
        csv_files = self.find_csv_files(period)

        if not csv_files:
            print(f"❌ ไม่พบไฟล์ CSV สำหรับงวด {period}")
            return False

        print(f"✓ พบไฟล์ CSV {len(csv_files)} ไฟล์")
        for sheet_name, csv_path in csv_files.items():
            print(f"  - {os.path.basename(csv_path)} → {sheet_name}")

        # สร้าง workbook ใหม่
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # ลบ sheet default

        # สร้าง sheet แต่ละอัน
        for sheet_name, csv_path in csv_files.items():
            print(f"\n⏳ กำลังประมวลผล: {sheet_name}...")

            # อ่านข้อมูลจาก CSV
            sheet_data = self.read_csv_data(csv_path)

            # สร้าง sheet
            ws = wb.create_sheet(title=sheet_name)

            # จัดรูปแบบ
            self.create_formatted_sheet(ws, sheet_data, period)

            print(f"  ✓ สำเร็จ")

        # สร้างชื่อไฟล์ output
        if not output_filename:
            # Parse period to get month and year
            month_num = period[:2] if len(period) >= 2 else "00"
            year_suffix = period[2:4] if len(period) >= 4 else "00"
            output_filename = f"Report_NT_BU_สะสม_{year_suffix}{month_num}.xlsx"

        output_path = os.path.join(self.base_path, output_filename)

        # บันทึกไฟล์
        wb.save(output_path)
        print(f"\n✅ สร้างไฟล์สำเร็จ: {output_filename}")
        print(f"   ตำแหน่ง: {output_path}")

        return True


def main():
    """ฟังก์ชันหลักสำหรับเรียกใช้งาน"""
    import sys

    # ตัวอย่างการใช้งาน
    if len(sys.argv) > 1:
        period = sys.argv[1]
    else:
        # ค้นหาไฟล์ CSV ที่มีอยู่และดึง period ออกมา
        for filename in os.listdir('.'):
            if filename.endswith('.csv') and ('ต้นทุน' in filename or 'บัญชี' in filename):
                # Extract period from filename (e.g., "1025" from "001_ต้นทุน_BU_1025 - 10-11-68.csv")
                match = re.search(r'_(\d{4})\s*-', filename)
                if match:
                    period = match.group(1)
                    print(f"🔍 ตรวจพบงวด: {period} จากไฟล์ {filename}")
                    break
        else:
            print("❌ ไม่พบไฟล์ CSV หรือไม่สามารถระบุงวดได้")
            print("\nวิธีใช้งาน:")
            print("  python csv_to_excel.py [period]")
            print("\nตัวอย่าง:")
            print("  python csv_to_excel.py 1025")
            return

    # สร้าง converter และแปลงไฟล์
    converter = CSVToExcelConverter(base_path='.')
    converter.convert(period)


if __name__ == "__main__":
    main()
