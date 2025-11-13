import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReporter:
    def __init__(self, output_file):
        self.writer = pd.ExcelWriter(output_file, engine='openpyxl')
        print(f"[Reporter]: Initialized for file: {output_file}")
        
        # กำหนด Style สีต่างๆ
        self.styles = {
            "High_Spike": PatternFill(start_color="FFC7CE", fill_type="solid"),
            "Spike_vs_Constant": PatternFill(start_color="FFC7CE", fill_type="solid"),
            "Low_Spike": PatternFill(start_color="FFEB9C", fill_type="solid"),
            "New_Item": PatternFill(start_color="C6E0B4", fill_type="solid"),
            "Negative_Value": PatternFill(start_color="FF0000", fill_type="solid"),
            "Low_Drop": PatternFill(start_color="FFEB9C", fill_type="solid"),
        }
        self.font_negative = Font(color="FFFFFF", bold=True)
        self.font_bold = Font(bold=True)
        self.align_right = Alignment(horizontal='right')
        self.num_format = "#,##0.00"

    def _compute_cell_anomaly(self, value, history, min_history=3):
        """
        คำนวณสถานะ anomaly ของ cell เดียว (เหมือน logic ใน CrosstabGenerator._get_status_helper)
        
        Parameters:
        - value: ค่าของเดือนปัจจุบัน
        - history: list ของค่าในอดีต (เรียงจากเก่า→ใหม่)
        - min_history: จำนวนข้อมูลอดีตขั้นต่ำ
        
        Returns: "Negative_Value" | "High_Spike" | "Low_Spike" | "New_Item" | "Normal"
        """
        # Case 1: ค่าติดลบ
        if value < 0:
            return "Negative_Value"
        
        # Case 2: ข้อมูลอดีตไม่พอ
        history_clean = [h for h in history if h > 0]
        if len(history_clean) < min_history:
            return "New_Item" if value > 0 else "Normal"
        
        # Case 3: คำนวณ % การเปลี่ยนแปลง
        avg_historical = np.mean(history_clean)
        
        if avg_historical == 0:
            pct_change = 0
        else:
            pct_change = abs((value - avg_historical) / avg_historical)
        
        # ถ้าเปลี่ยนน้อยกว่า 10% → Normal
        if pct_change < 0.10:
            return "Normal"
        
        # Case 4: คำนวณ IQR
        Q1 = np.percentile(history_clean, 25)
        Q3 = np.percentile(history_clean, 75)
        IQR = Q3 - Q1
        
        # ถ้า IQR = 0 (ข้อมูลคงที่)
        if IQR == 0:
            if pct_change < 0.15:
                return "Normal"
            if Q1 == 0 and value > 0:
                return "High_Spike"
            if value != Q1:
                return "Spike_vs_Constant"
            return "Normal"
        
        # Case 5: ใช้ IQR Method
        k = 2.0  # Sensitivity factor
        lower_fence = max(0, Q1 - (k * IQR))
        upper_fence = Q3 + (k * IQR)
        
        if value > upper_fence:
            return "High_Spike"
        if value < lower_fence:
            return "Low_Spike"
        
        return "Normal"

    def _build_anomaly_map(self, df_report, date_cols_sorted, min_history=3):
        """
        สร้าง anomaly map สำหรับทุก cell ในตาราง Crosstab
        
        Returns: dict[(row_idx, col_name)] = "Negative_Value" | "High_Spike" | ...
        """
        print("[Reporter]:    Computing anomaly status for all cells...")
        
        anomaly_map = {}
        
        # Loop ทุกแถวใน dataframe
        for row_idx, row in df_report.iterrows():
            # ดึงค่าทุกเดือนออกมา
            values = [row[col] for col in date_cols_sorted]
            
            # คำนวณ anomaly สำหรับแต่ละเดือน
            for i, col_name in enumerate(date_cols_sorted):
                current_value = values[i]
                
                # ประวัติ = เดือนก่อนหน้า (ไม่รวมเดือนปัจจุบัน)
                if i == 0:
                    history = []  # เดือนแรก ไม่มีประวัติ
                else:
                    history = values[:i]  # เอาทุกเดือนก่อนหน้า
                
                # คำนวณสถานะ
                status = self._compute_cell_anomaly(current_value, history, min_history)
                
                # เก็บเฉพาะที่ไม่ใช่ Normal
                if status not in ["Normal", "New_Item"]:
                    anomaly_map[(row_idx, col_name)] = status
        
        print(f"[Reporter]:    ✓ Found {len(anomaly_map)} anomalies across all cells")
        return anomaly_map

    def _add_legend(self, ws):
        """สร้างตารางคำอธิบายสี (Legend) ต่อท้ายข้อมูล"""
        start_row = ws.max_row + 4
        
        cell_header = ws.cell(row=start_row, column=2, value="คำอธิบายความหมายสี (Color Legend)")
        cell_header.font = self.font_bold
        
        legend_data = [
            ("High_Spike",      "ยอดพุ่งสูงผิดปกติ (High Spike)"),
            ("Low_Spike",       "ยอดตกลงต่ำผิดปกติ (Low Drop)"),
            ("Negative_Value",  "ยอดติดลบ")
        ]

        for i, (key, desc) in enumerate(legend_data):
            r = start_row + 1 + i
            
            c_color = ws.cell(row=r, column=2, value="     ")
            if key in self.styles:
                c_color.fill = self.styles[key]
                if key == "Negative_Value": 
                    c_color.font = self.font_negative
            
            c_desc = ws.cell(row=r, column=3, value=desc)
            c_desc.alignment = Alignment(horizontal='left')

        print(f"[Reporter]:    ✓ Added Color Legend at row {start_row}")

    def add_crosstab_sheet(self, df_report, df_anomaly_log, dimensions, date_col_name, date_cols_sorted):
        """เพิ่ม Crosstab Sheet และทาสีตาม anomaly ที่คำนวณจากข้อมูล Crosstab โดยตรง"""
        if df_report.empty: 
            return

        print("[Reporter]: Adding Crosstab Sheet with Cell Highlighting...")
        
        # ✅ คำนวณ anomaly สำหรับทุก cell
        anomaly_map = self._build_anomaly_map(df_report, date_cols_sorted, min_history=3)
        
        # เขียน DataFrame ลง Excel
        sheet_name = 'Crosstab_Report'
        df_report.to_excel(self.writer, sheet_name=sheet_name, index=False)
        
        ws = self.writer.sheets[sheet_name]
        
        header_cells = ws[1]
        col_map = {cell.value: (cell.column, cell.column_letter) for cell in header_cells}
        
        # Loop ทุกแถวเพื่อ format และทาสี
        for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
            df_row_idx = excel_row_idx - 2  # แปลง Excel row → DataFrame row index
            
            for col_name, (col_idx, col_letter) in col_map.items():
                cell = ws[f"{col_letter}{excel_row_idx}"]
                
                if col_name in date_cols_sorted:
                    # Format ตัวเลข
                    cell.number_format = self.num_format
                    cell.alignment = self.align_right
                    
                    # ✅ ทาสีตาม anomaly_map
                    anomaly_status = anomaly_map.get((df_row_idx, col_name))
                    if anomaly_status:
                        if anomaly_status in self.styles:
                            cell.fill = self.styles[anomaly_status]
                            if anomaly_status == "Negative_Value":
                                cell.font = self.font_negative
                
                elif col_name == 'ANOMALY_STATUS':
                    # ทาสีตามค่าใน column
                    status = cell.value
                    if status in self.styles:
                        cell.fill = self.styles[status]
                        if status == "Negative_Value":
                            cell.font = self.font_negative
                
                elif col_name == 'PCT_CHANGE':
                    cell.number_format = '0.00"%"'
                    cell.alignment = self.align_right
                elif col_name in ['LATEST_VALUE', 'AVG_HISTORICAL']:
                    cell.number_format = self.num_format
                    cell.alignment = self.align_right

        # จัดความกว้างคอลัมน์
        for col_name, (idx, letter) in col_map.items():
            if col_name in dimensions:
                ws.column_dimensions[letter].width = 30
            elif col_name in date_cols_sorted:
                ws.column_dimensions[letter].width = 15
            elif col_name == 'ANOMALY_STATUS':
                ws.column_dimensions[letter].width = 20
            else:
                ws.column_dimensions[letter].width = 18

        ws.freeze_panes = f'{get_column_letter(len(dimensions) + 1)}2'

        # เพิ่ม Legend
        self._add_legend(ws)
        
        print(f"[Reporter]:    ✓ Crosstab sheet created with accurate cell-by-cell highlighting")

    def add_audit_log_sheet(self, df_log, sheet_name, cols_to_show):
        print(f"[Reporter]: Adding Log Sheet: {sheet_name}...")
        if df_log.empty:
            df_log = pd.DataFrame({'Message': ['No Anomalies Found']})
            cols_to_show = ['Message']
        valid_cols = [c for c in cols_to_show if c in df_log.columns]
        df_log[valid_cols].to_excel(self.writer, sheet_name=sheet_name, index=False)
        ws = self.writer.sheets[sheet_name]
        for i, col in enumerate(valid_cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = 25

    def save(self):
        try:
            self.writer.close()
            print(f"[Reporter]: ✓ Report saved: {self.writer.path}")
        except:
            print(f"[Reporter]: ✓ Report saved.")