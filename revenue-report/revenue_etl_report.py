import pandas as pd
import numpy as np
import glob
import os
from pathlib import Path
from datetime import datetime
import platform
from typing import Dict, Any
from revenue_reconciliation import RevenueReconciliation, ReconciliationError
from logger_utils import ETLLogger


# ============================================================================
# CONFIG ADAPTER - แปลง dict config เป็น object-like access
# ============================================================================
class ConfigAdapter:
    """
    Adapter class สำหรับแปลง config dictionary เป็น attribute-based access
    เพื่อให้โค้ดเดิมใช้งานได้โดยไม่ต้องแก้ไขมาก
    """
    def __init__(self, config_dict: Dict[str, Any]):
        """
        Args:
            config_dict: Configuration dictionary จาก ConfigManager
        """
        # ETL Module config
        etl_config = config_dict

        # ปีที่ประมวลผล
        self.YEAR = etl_config['year']

        # Month Settings (รองรับ month filtering)
        self.end_month = etl_config.get('end_month', None)  # เดือนสุดท้ายที่ต้องการประมวลผล
        self.fi_month = etl_config.get('fi_month', None)    # เดือนของ FI file

        # Excel Formatting Settings (เพิ่มใหม่)
        excel_format = etl_config.get('excel_format', {})
        self.EXCEL_FONT_NAME = excel_format.get('font_name', 'TH Sarabun New')
        self.EXCEL_FONT_SIZE = excel_format.get('font_size', 16)
        self.EXCEL_HEADER_SIZE = excel_format.get('header_size', 16)

        # Reconciliation Settings
        reconcile = etl_config.get('reconciliation', {})
        self.RECONCILE_FI_MONTH = reconcile.get('fi_month', '10')
        self.RECONCILE_TOLERANCE = reconcile.get('tolerance', 0.00)
        self.ENABLE_RECONCILIATION = reconcile.get('enabled', True)

        # Master Files
        master_files = etl_config.get('master_files', {})
        self.MASTER_PRODUCT_FILE = master_files.get('product', '')
        self.MASTER_GL_FILE = master_files.get('gl_code', '')
        self.MAPPING_CC_FILE = master_files.get('mapping_cc', '')
        self.MAPPING_PRODUCT_FILE = master_files.get('mapping_product', '')

        # Input File Patterns (ใช้ default ถ้าไม่มีใน config)
        self.INPUT_FILE_PATTERNS = etl_config.get('input_file_patterns', [
            "TRN_REVENUE_NT1_*.csv",
            "TRN_REVENUE_ADJ_GL_NT1_*.csv"
        ])
        self.ADJ_MONTHLY_PATTERN = etl_config.get('adj_monthly_pattern', "TRN_REVENUE_ADJ_*.csv")
        self.ADJ_YTD_PATTERN = etl_config.get('adj_ytd_pattern', "TRN_REVENUE_ADJ_YTD_*.csv")

        # Business Rules
        business_rules = etl_config.get('business_rules', {})
        self.EXCLUDE_BUSINESS_GROUP = business_rules.get('exclude_business_group', 'รายได้อื่น')
        self.NON_TELECOM_SERVICE_GROUP = business_rules.get('non_telecom_service_group', 'กลุ่มบริการอื่นไม่ใช่โทรคมนาคม')
        self.NEW_ADJ_BUSINESS_GROUP = business_rules.get('new_adj_business_group', 'ผลตอบแทนทางการเงินและรายได้อื่น')
        self.FINANCIAL_INCOME_NAME = business_rules.get('financial_income_name', 'ผลตอบแทนทางการเงิน')
        self.OTHER_REVENUE_ADJ_NAME = business_rules.get('other_revenue_adj_name', 'รายได้อื่น')

        # Output Files
        output_files = etl_config.get('output_files', {})
        self.OUTPUT_CONCAT_FILE = output_files.get('concat', f"trn_revenue_nt_{self.YEAR}.csv")
        self.OUTPUT_MAPPED_CC_FILE = output_files.get('mapped_cc', f"revenue_new_cc_{self.YEAR}.csv")
        self.OUTPUT_MAPPED_PRODUCT_FILE = output_files.get('mapped_product', f"revenue_mapped_product_{self.YEAR}_.csv")
        self.OUTPUT_FINAL_REPORT_FILE = output_files.get('final_report', f"REVENUE_NT_REPORT_{self.YEAR}.csv")
        self.ERROR_GL_FILE = output_files.get('error_gl', f"error_gl_REVENUE_NT_REPORT_{self.YEAR}.csv")
        self.ERROR_PRODUCT_FILE = output_files.get('error_product', f"error_product_REVENUE_NT_REPORT_{self.YEAR}.csv")

        # Required Columns
        self.REQUIRED_COLUMNS = etl_config.get('required_columns', [
            "YEAR", "MONTH", "CUSTOMER_GROUP_KEY", "PRODUCT_KEY",
            "SUB_PRODUCT_KEY", "GL_CODE", "COST_CENTER", "REVENUE_VALUE"
        ])

        # Special Mappings
        self.SPECIAL_MAPPINGS = etl_config.get('special_mappings', [
            {
                "name": "GSaaS to Other Revenue",
                "condition": {
                    "PRODUCT_KEY": "102010407",
                    "GL_CODE": "46400101"
                },
                "mapping": {
                    "PRODUCT_KEY": "292020407",
                    "SUB_PRODUCT_KEY": "1"
                }
            }
        ])

        # Validation Thresholds
        self.GRAND_TOTAL_DIFF_THRESHOLD = etl_config.get('grand_total_diff_threshold', 0.01)

        # Anomaly Detection Settings
        anomaly = etl_config.get('anomaly_detection', {})
        self.ANOMALY_IQR_MULTIPLIER = anomaly.get('iqr_multiplier', 1.5)
        self.ANOMALY_MIN_HISTORY = anomaly.get('min_history', 3)
        self.ANOMALY_ROLLING_WINDOW = anomaly.get('rolling_window', 6)
        self.ENABLE_HISTORICAL_HIGHLIGHT = anomaly.get('enable_historical_highlight', True)

        # Anomaly Levels
        self.ANOMALY_LEVELS = {
            "product": {"group_by": ["BUSINESS_GROUP", "SERVICE_GROUP", "PRODUCT_KEY", "PRODUCT_NAME"]},
            "service": {"group_by": ["BUSINESS_GROUP", "SERVICE_GROUP"]},
            "business": {"group_by": ["BUSINESS_GROUP"]},
            "grand_total": {"group_by": []}
        }

    def get_paths(self):
        """
        Dummy method for backward compatibility
        จะไม่ถูกใช้งานเพราะ paths จะถูกส่งมาจาก ConfigManager แล้ว
        """
        raise NotImplementedError("get_paths() should not be called when using ConfigAdapter")


# ============================================================================
# LEGACY CONFIGURATION - สำหรับ backward compatibility เท่านั้น (ไม่ควรใช้แล้ว)
# ============================================================================
class Config:
    """
    Configuration สำหรับ Revenue ETL Pipeline
    แก้ไขค่า config ทั้งหมดที่นี่
    """
    
    # ============ ปี ที่ต้องการประมวลผล ============
    YEAR = "2025"

    # ============ Reconciliation Settings (เพิ่มใหม่) ============
    RECONCILE_FI_MONTH = "10"       # เดือนของไฟล์ FI
    RECONCILE_TOLERANCE = 0.00      # ยอดความแตกต่างที่ยอมรับได้
    ENABLE_RECONCILIATION = True    # True = เปิด, False = ปิด
    
    # ============ ชื่อไฟล์ Master ============
    MASTER_PRODUCT_FILE = "MASTER_PRODUCT_NT_2025.csv"
    MASTER_GL_FILE = "source/MASTER_REVENUE_GL_CODE_NT1_NT_20250723.csv"
    MAPPING_CC_FILE = "MAPPING_CC.csv"
    MAPPING_PRODUCT_FILE = "clean/MAP_PRODUCT_NT_NEW_2024.csv"
    
    # ============ Pattern สำหรับหาไฟล์ต้นทาง ============
    # 1. ไฟล์หลัก NT1
    INPUT_FILE_PATTERNS = [
        "TRN_REVENUE_NT1_*.csv",
        "TRN_REVENUE_ADJ_GL_NT1_*.csv"
    ]
    
    # 2. ไฟล์ ADJ (ผลตอบแทนทางการเงิน, รายได้อื่น)
    ADJ_MONTHLY_PATTERN = "TRN_REVENUE_ADJ_*.csv"
    ADJ_YTD_PATTERN = "TRN_REVENUE_ADJ_YTD_*.csv"

    # ============ ค่าคงที่สำหรับกรองและสร้างรายงาน ============
    EXCLUDE_BUSINESS_GROUP = "รายได้อื่น" # จาก Master Product ที่จะกรองออก
    NON_TELECOM_SERVICE_GROUP = "กลุ่มบริการอื่นไม่ใช่โทรคมนาคม"
    
    NEW_ADJ_BUSINESS_GROUP = "ผลตอบแทนทางการเงินและรายได้อื่น"
    FINANCIAL_INCOME_NAME = "ผลตอบแทนทางการเงิน"
    OTHER_REVENUE_ADJ_NAME = "รายได้อื่น" # จาก ADJ File
    
    
    # ============ ชื่อไฟล์ Output ============
    OUTPUT_CONCAT_FILE = f"trn_revenue_nt_{YEAR}.csv"
    OUTPUT_MAPPED_CC_FILE = f"revenue_new_cc_{YEAR}.csv"
    OUTPUT_MAPPED_PRODUCT_FILE = f"revenue_mapped_product_{YEAR}_.csv"
    OUTPUT_FINAL_REPORT_FILE = f"REVENUE_NT_REPORT_{YEAR}.csv"
    
    # ============ ชื่อไฟล์ Error Log ============
    ERROR_GL_FILE = f"error_gl_REVENUE_NT_REPORT_{YEAR}.csv"
    ERROR_PRODUCT_FILE = f"error_product_REVENUE_NT_REPORT_{YEAR}.csv"
    
    # ============ Path Configuration ============
    @staticmethod
    def get_paths():
        """
        กำหนด path ตามระบบปฏิบัติการ
        Returns: dict ของ paths ทั้งหมด
        """
        os_platform = platform.system()
        
        if os_platform == "Darwin":  # macOS
            base_path = "/Users/seal/Library/CloudStorage/OneDrive-Personal/share/Datasource"
            master_path = "/Users/seal/Library/CloudStorage/OneDrive-Personal/share/master"
        elif os_platform == "Linux":
            base_path = "/home/seal/nt/data/2025"
            master_path = "/home/seal/nt/master"
        elif os_platform == "Windows":
            base_path = r"C:\Users\00320845\OneDrive\share\Datasource"
            master_path = r"C:\Users\00320845\OneDrive\share\master"
        else:
            raise ValueError(f"ไม่รองรับระบบปฏิบัติการ: {os_platform}")
        
        return {
            "base": base_path,
            "master": master_path,
            "input": os.path.join(base_path, Config.YEAR, "revenue"),
            "output": os.path.join(base_path, Config.YEAR, "revenue", "output"),
            "final_output": os.path.join(base_path, "all", "revenue", Config.YEAR)
        }
    
    # ============ Column Names ============
    REQUIRED_COLUMNS = [
        "YEAR", "MONTH", "CUSTOMER_GROUP_KEY", "PRODUCT_KEY",
        "SUB_PRODUCT_KEY", "GL_CODE", "COST_CENTER", "REVENUE_VALUE"
    ]
    
    # ============ Special Mapping Rules ============
    SPECIAL_MAPPINGS = [
        {
            "name": "GSaaS to Other Revenue",
            "condition": {
                "PRODUCT_KEY": "102010407",
                "GL_CODE": "46400101"
            },
            "mapping": {
                "PRODUCT_KEY": "292020407",
                "SUB_PRODUCT_KEY": "1"
            }
        }
    ]
    
    # ============ Validation Thresholds ============
    GRAND_TOTAL_DIFF_THRESHOLD = 0.01  # บาท - สำหรับ rounding error

    # ============ Anomaly Detection Settings ============
    ANOMALY_IQR_MULTIPLIER = 1.5  # ค่า k สำหรับ Tukey's Fences
    ANOMALY_MIN_HISTORY = 3  # จำนวนเดือนขั้นต่ำที่ต้องมีเพื่อตรวจสอบ

    # [เพิ่มบรรทัดนี้] จำนวนเดือนสำหรับ Rolling Average
    ANOMALY_ROLLING_WINDOW = 6

    # True = ตรวจสอบและระบายสีทุกเดือน, False = ทำเหมือนเดิม (เฉพาะ Report แยก)
    ENABLE_HISTORICAL_HIGHLIGHT = True
    
    # ระดับการตรวจสอบ
    ANOMALY_LEVELS = {
        "product": {"group_by": ["BUSINESS_GROUP", "SERVICE_GROUP", "PRODUCT_KEY", "PRODUCT_NAME"]},
        "service": {"group_by": ["BUSINESS_GROUP", "SERVICE_GROUP"]},
        "business": {"group_by": ["BUSINESS_GROUP"]},
        "grand_total": {"group_by": []}
    }


# ============================================================================
# ETL Pipeline Class
# ============================================================================
class RevenueETL:
    """
    ETL Pipeline สำหรับประมวลผลข้อมูลรายได้
    ขั้นตอน:
    1. รวมไฟล์ CSV ต้นทาง (TRN_REVENUE_NT1_*.csv และ TRN_REVENUE_ADJ_GL_NT1_*.csv)
    2. Mapping Cost Center
    3. Mapping Product & Sub Product
    4. Merge กับ Master Files, กรอง, รวม ADJ Data, และสร้างรายงานสุดท้าย
    """
    
    def __init__(self, config=None, paths=None):
        """
        Args:
            config: Configuration dictionary (from ConfigManager) หรือ Config class (legacy) หรือ None
            paths: Dictionary ของ paths (จาก ConfigManager) หรือ None
        """
        # สำหรับ V2: รับ config dict และ paths dict จาก ConfigManager
        if isinstance(config, dict):
            self.config_dict = config  # เก็บ dict ไว้สำหรับส่งต่อ
            self.config = ConfigAdapter(config)  # แปลงเป็น object-like
            self.paths = paths if paths else config.get('paths', {})
        # สำหรับ V1 (legacy): รับ Config class
        else:
            self.config_dict = None
            self.config = config or Config
            self.paths = self.config.get_paths()

        # Setup logger
        self.logger = ETLLogger.get_logger('etl_module')

        self.setup_directories()

        # ตัวแปรสำหรับเก็บ YTD data ไว้ใช้ใน report
        self.df_adj_ytd = pd.DataFrame() 

    def read_csv_robust(self, file_path, **kwargs):
        """
        [NEW] Reads a CSV file with robust encoding handling.
        Tries encodings in order: utf-8 -> tis-620 -> cp1252 -> latin1
        """
        encodings = ['utf-8', 'tis-620', 'cp874', 'cp1252', 'latin1']
        
        # Remove 'encoding' from kwargs if present to avoid conflict
        kwargs.pop('encoding', None)

        for encoding in encodings:
            try:
                # self.log(f"  Attempting read with {encoding}...")
                df = pd.read_csv(file_path, encoding=encoding, **kwargs)
                if encoding != 'utf-8':
                     self.log(f"  ℹ️  Read success with encoding: {encoding}")
                return df
            except (UnicodeDecodeError, LookupError):
                continue
            except Exception as e:
                # If it's a structural error (not encoding), raise immediately
                self.log(f"  ❌ Read Error ({encoding}): {e}")
                raise e
        
        raise UnicodeDecodeError(f"Failed to read {os.path.basename(file_path)} with any of {encodings}")
 
        
    def setup_directories(self):
        """สร้าง directory ถ้ายังไม่มี"""
        Path(self.paths["output"]).mkdir(parents=True, exist_ok=True)
        Path(self.paths["final_output"]).mkdir(parents=True, exist_ok=True)

        self.logger.info("=" * 80)
        self.logger.info("PATH CONFIGURATION")
        self.logger.info("=" * 80)
        for key, path in self.paths.items():
            self.logger.info(f"{key:15s}: {path}")
        self.logger.info("=" * 80)
        
    def log(self, message, grand_total=None):
        """แสดงข้อความ log พร้อม timestamp"""
        # [FIX] ตรวจสอบประเภทของ grand_total
        # ถ้าเป็น string, ถือว่าเป็น level (เช่น "ERROR", "SUCCESS")
        # ถ้าเป็นตัวเลข (int/float), ถือว่าเป็นยอดรวม

        if isinstance(grand_total, str):
            # grand_total เป็น level
            self.logger.log(message, grand_total)
        elif isinstance(grand_total, (int, float)):
            # grand_total เป็นตัวเลข
            self.logger.info(message)
            self.logger.info(f"Grand Total = {grand_total:,.2f}")
        else:
            # กรณี grand_total=None หรืออื่นๆ
            self.logger.info(message)

    def step0_reconcile_revenue(self):
        """ตรวจสอบความถูกต้องของข้อมูล (Reconciliation)"""
        self.log("=" * 80)
        self.log("STEP 0: RECONCILIATION - ตรวจสอบความถูกต้องของข้อมูล")
        self.log("=" * 80)
        
        if not self.config.ENABLE_RECONCILIATION:
            self.log("⚠️  Reconciliation ถูกปิดใช้งาน")
            return None

        # สำหรับ V2: ส่ง config_dict (ถ้ามี), สำหรับ V1: ส่ง config object
        config_to_pass = self.config_dict if self.config_dict is not None else self.config
        reconciler = RevenueReconciliation(config_to_pass, self.paths)
        
        fi_file = os.path.join(
            self.paths['base'], 
            self.config.YEAR, 
            'fi', 
            'output', 
            f'pl_revenue_nt_output_{self.config.YEAR}{self.config.RECONCILE_FI_MONTH}.csv'
        )
        
        trn_file = os.path.join(
            self.paths['output'], 
            self.config.OUTPUT_CONCAT_FILE
        )
        
        self.log(f"ไฟล์ FI: {os.path.basename(fi_file)}")
        self.log(f"ไฟล์ TRN: {os.path.basename(trn_file)}")
        
        try:
            result = reconciler.reconcile_revenue(
                fi_file_path=fi_file,
                trn_file_path=trn_file,
                tolerance=self.config.RECONCILE_TOLERANCE
            )
            
            self.log("=" * 80)
            self.log("✓ Reconciliation สำเร็จ!", "SUCCESS")
            self.log("=" * 80)
            return result
            
        except ReconciliationError as e:
            self.log("=" * 80)
            self.log("❌ Reconciliation ล้มเหลว - หยุดการทำงาน", "ERROR")
            self.log("=" * 80)
            # Re-raise แต่ใช้ raise from None เพื่อลด traceback
            raise ReconciliationError(str(e), e.reconcile_results) from None
        
        except FileNotFoundError as e:
            self.log(f"❌ ไม่พบไฟล์: {e}", "ERROR")
            self.log("💡 กรุณาตรวจสอบ:")
            self.log(f"   1. ไฟล์ FI อยู่ที่: {fi_file}")
            self.log(f"   2. RECONCILE_FI_MONTH = '{self.config.RECONCILE_FI_MONTH}' ถูกต้องหรือไม่")
    
    def step1_concat_revenue_files(self):
        """
        ขั้นตอนที่ 1: รวมไฟล์ CSV ต้นทาง (NT1)
        - TRN_REVENUE_NT1_*.csv และ TRN_REVENUE_ADJ_GL_NT1_*.csv
        - จัดการ GL_CODE และ GL_CODE_NT1 ให้เป็น GL_CODE
        """
        self.log("=" * 80)
        self.log("STEP 1: รวมไฟล์ CSV ต้นทาง (NT1)")
        self.log("=" * 80)
        
        # หาไฟล์ทั้งหมดที่ต้องการ
        all_files = []
        for p in self.config.INPUT_FILE_PATTERNS:
            pattern = os.path.join(self.paths["input"], p)
            self.log(f"ค้นหา pattern: {pattern}")
            all_files.extend(glob.glob(pattern))
        
        # ใช้ set เพื่อลบไฟล์ที่อาจซ้ำกัน (ถ้ามี) แล้วเรียงลำดับ
        files = sorted(list(set(all_files)))
        
        if not files:
            patterns = [os.path.join(self.paths["input"], p) for p in self.config.INPUT_FILE_PATTERNS]
            raise FileNotFoundError(f"ไม่พบไฟล์ที่ตรงกับ patterns: {patterns}")
        
        self.log(f"พบไฟล์ทั้งหมด {len(files)} ไฟล์")
        
        df_combined = pd.DataFrame()
        
        for file in files:
            self.log(f"กำลังอ่านไฟล์: {os.path.basename(file)}")
            
            # อ่านไฟล์
            try:
                df = self.read_csv_robust(
                    file,
                    converters={
                        "YEAR": str,
                        "MONTH": int,
                        "COST_CENTER": str,
                        "CUSTOMER_GROUP_KEY": str,
                        "PRODUCT_KEY": str,
                        "SUB_PRODUCT_KEY": int
                    }
                )
            except Exception as e:
                self.log(f"  ❌ เกิดข้อผิดพลาดในการอ่านไฟล์ {file}: {e} - ข้ามไฟล์นี้")
                continue
            
            # จัดการชื่อคอลัมน์
            df.columns = df.columns.str.strip()
            
            # แสดง GL columns ที่พบ
            gl_columns = df.filter(like='GL').columns.tolist()
            self.log(f"  พบ GL columns: {gl_columns}")
            
            # จัดการ REVENUE_VALUE
            df = df.dropna(subset=["REVENUE_VALUE"])
            df["REVENUE_VALUE"] = df["REVENUE_VALUE"].astype(str)
            df["REVENUE_VALUE"] = df["REVENUE_VALUE"].str.replace(",", "", regex=False)
            df["REVENUE_VALUE"] = df["REVENUE_VALUE"].str.replace(r"\(", "-", regex=True)
            df["REVENUE_VALUE"] = df["REVENUE_VALUE"].str.replace(r"\)", "", regex=True)
            df["REVENUE_VALUE"] = df["REVENUE_VALUE"].str.replace(" ", "", regex=False)
            df["REVENUE_VALUE"] = pd.to_numeric(df["REVENUE_VALUE"], errors='coerce')
            
            # จัดการ GL_CODE และ GL_CODE_NT1
            if "GL_CODE_NT1" in df.columns:
                self.log(f"  ใช้ GL_CODE_NT1 เป็น GL_CODE")
                df["GL_CODE"] = df["GL_CODE_NT1"].astype(str)
            elif "GL_CODE" in df.columns:
                self.log(f"  ใช้ GL_CODE")
                df["GL_CODE"] = df["GL_CODE"].astype(str)
            else:
                if len(gl_columns) > 0:
                    self.log(f"  ใช้ {gl_columns[0]} เป็น GL_CODE")
                    df["GL_CODE"] = df[gl_columns[0]].astype(str)
                else:
                    self.log(f"  ⚠️  ไม่พบ column GL_CODE หรือ GL_CODE_NT1 ในไฟล์ {file} - สร้างคอลัมน์เปล่า")
                    df["GL_CODE"] = np.nan # หรือ "NA"
            
            # เลือกเฉพาะคอลัมน์ที่ต้องการ
            try:
                df = df[self.config.REQUIRED_COLUMNS]
            except KeyError as e:
                self.log(f"  ❌ ขาดคอลัมน์ที่จำเป็นใน {file}: {e} - ข้ามไฟล์นี้")
                continue

            # ลบแถวที่ PRODUCT_KEY เป็น null
            df = df.dropna(subset=["PRODUCT_KEY"])
            
            file_total = df["REVENUE_VALUE"].sum()
            self.log(f"  ยอดรวมของไฟล์นี้: {file_total:,.2f}")
            self.log(f"  จำนวนแถว: {len(df):,}")
            
            df_combined = pd.concat([df_combined, df], ignore_index=True)
        
        # แปลง MONTH เป็น string แบบ 2 หลัก
        df_combined["MONTH"] = df_combined["MONTH"].astype(int).astype(str).str.zfill(2)
        df_combined["SUB_PRODUCT_KEY"] = df_combined["SUB_PRODUCT_KEY"].astype(int).astype(str)

        # === [NEW] Month Filtering: กรองข้อมูลตามช่วงเดือนที่กำหนด ===
        # ถ้ามี end_month ใน config ให้กรองเฉพาะเดือนที่ต้องการ
        if hasattr(self.config, 'end_month') and self.config.end_month:
            end_month = self.config.end_month
            year = self.config.YEAR

            # แปลง MONTH กลับเป็น int สำหรับการเปรียบเทียบ
            df_combined_temp = df_combined.copy()
            df_combined_temp['MONTH_INT'] = df_combined_temp['MONTH'].astype(int)

            # Filter: เอาเฉพาะปีและเดือนที่ต้องการ
            before_filter_count = len(df_combined_temp)
            before_filter_total = df_combined_temp["REVENUE_VALUE"].sum()

            df_combined = df_combined_temp[
                (df_combined_temp['YEAR'] == year) &
                (df_combined_temp['MONTH_INT'] <= end_month)
            ].copy()

            # ลบ MONTH_INT ออก (ใช้แค่ชั่วคราว)
            df_combined = df_combined.drop(columns=['MONTH_INT'])

            after_filter_count = len(df_combined)
            after_filter_total = df_combined["REVENUE_VALUE"].sum()

            self.log("=" * 80)
            self.log(f"📅 MONTH FILTERING: กรองข้อมูลเดือน 1 - {end_month:02d}/{year}")
            self.log(f"  ก่อนกรอง: {before_filter_count:,} แถว, ยอดรวม: {before_filter_total:,.2f}")
            self.log(f"  หลังกรอง: {after_filter_count:,} แถว, ยอดรวม: {after_filter_total:,.2f}")
            self.log(f"  ลดลง: {before_filter_count - after_filter_count:,} แถว")
            self.log("=" * 80)
        else:
            self.log("ℹ️  Month Filtering: ไม่กำหนด - ประมวลผลทุกเดือน")
        # === [END] Month Filtering ===

        # ตรวจสอบ GL_CODE ที่ได้
        unique_gl = df_combined["GL_CODE"].nunique()
        self.log(f"จำนวน GL_CODE ที่ไม่ซ้ำกัน: {unique_gl}")
        
        # บันทึกไฟล์
        output_file = os.path.join(self.paths["output"], self.config.OUTPUT_CONCAT_FILE)
        df_combined.to_csv(output_file, index=False, float_format="%.2f")
        
        self.log(f"บันทึกไฟล์: {output_file}")
        self.log(f"รวมไฟล์เสร็จสิ้น - จำนวนแถวทั้งหมด: {len(df_combined):,}")
        self.log(f"รวมไฟล์เสร็จสิ้น", df_combined["REVENUE_VALUE"].sum())
        
        return df_combined
    
    def step2_mapping_cost_center(self, df):
        """ขั้นตอนที่ 2: Mapping Cost Center"""
        self.log("=" * 80)
        self.log("STEP 2: Mapping Cost Center")
        self.log("=" * 80)
        
        # อ่านไฟล์ mapping
        mapping_file = os.path.join(self.paths["master"], self.config.MAPPING_CC_FILE)
        
        if not os.path.exists(mapping_file):
            self.log(f"ไม่พบไฟล์ {mapping_file} - ข้ามขั้นตอนนี้")
            return df
        
        df_mapping = pd.read_csv(mapping_file)
        map_dict = dict(df_mapping.values)
        
        self.log(f"พบการ mapping {len(map_dict)} รายการ")
        
        # Mapping cost center
        original_cc = df["COST_CENTER"].copy()
        df["COST_CENTER"] = df["COST_CENTER"].map(map_dict).fillna(df["COST_CENTER"])
        
        # นับจำนวนที่ถูก map
        mapped_count = (original_cc != df["COST_CENTER"]).sum()
        self.log(f"จำนวน Cost Center ที่ถูก mapping: {mapped_count:,}")
        
        # บันทึกไฟล์
        output_file = os.path.join(self.paths["output"], self.config.OUTPUT_MAPPED_CC_FILE)
        df.to_csv(output_file, index=False, float_format="%.2f")
        
        self.log(f"บันทึกไฟล์: {output_file}")
        self.log(f"Mapping Cost Center เสร็จสิ้น", df["REVENUE_VALUE"].sum())
        
        return df
    
    def step3_mapping_product(self, df):
        """ขั้นตอนที่ 3: Mapping Product & Sub Product"""
        self.log("=" * 80)
        self.log("STEP 3: Mapping Product & Sub Product")
        self.log("=" * 80)
        
        # อ่านไฟล์ mapping product
        mapping_file = os.path.join(self.paths["master"], self.config.MAPPING_PRODUCT_FILE)
        
        if not os.path.exists(mapping_file):
            self.log(f"ไม่พบไฟล์ {mapping_file} - ข้ามขั้นตอนนี้")
            return df
        
        df_mapping = pd.read_csv(
            mapping_file,
            converters={
                "PRODUCT_KEY_OLD": str,
                "SUB_PRODUCT_KEY_OLD": str,
                "PRODUCT_KEY": str,
                "SUB_PRODUCT_KEY": str,
                "GL_CODE": str
            }
        )
        
        # เลือกเฉพาะคอลัมน์ที่ต้องการ
        df_mapping = df_mapping[["PRODUCT_KEY_OLD", "SUB_PRODUCT_KEY_OLD", 
                                  "PRODUCT_KEY", "SUB_PRODUCT_KEY"]]
        
        # แปลงเป็น int แล้วกลับเป็น str (เพื่อตัด 0 ข้างหน้า)
        df_mapping["SUB_PRODUCT_KEY_OLD"] = df_mapping["SUB_PRODUCT_KEY_OLD"].astype(int).astype(str)
        df_mapping["SUB_PRODUCT_KEY"] = df_mapping["SUB_PRODUCT_KEY"].astype(int).astype(str)
        
        # สร้าง composite key
        df["product_key_sub_product"] = df["PRODUCT_KEY"] + df["SUB_PRODUCT_KEY"]
        df_mapping["product_key_sub_product"] = (df_mapping["PRODUCT_KEY_OLD"] + 
                                                  df_mapping["SUB_PRODUCT_KEY_OLD"])
        
        self.log(f"พบการ mapping product {len(df_mapping)} รายการ")
        
        # Mapping PRODUCT_KEY
        map_product = dict(zip(df_mapping["product_key_sub_product"], 
                              df_mapping["PRODUCT_KEY"]))
        original_product = df["PRODUCT_KEY"].copy()
        df["PRODUCT_KEY"] = df["product_key_sub_product"].map(map_product).fillna(df["PRODUCT_KEY"])
        
        # Mapping SUB_PRODUCT_KEY
        map_sub_product = dict(zip(df_mapping["product_key_sub_product"], 
                                   df_mapping["SUB_PRODUCT_KEY"]))
        df["SUB_PRODUCT_KEY"] = df["product_key_sub_product"].map(map_sub_product).fillna(df["SUB_PRODUCT_KEY"])
        
        # นับจำนวนที่ถูก map
        mapped_count = (original_product != df["PRODUCT_KEY"]).sum()
        self.log(f"จำนวน Product ที่ถูก mapping: {mapped_count:,}")
        
        # ลบคอลัมน์ชั่วคราว
        df = df.drop(columns=["product_key_sub_product"])
        
        # Apply special mapping rules
        self.log("กำลัง apply special mapping rules...")
        for rule in self.config.SPECIAL_MAPPINGS:
            self.log(f"  - {rule['name']}")

            # สร้าง condition
            condition = pd.Series([True] * len(df))
            for cond_col, cond_val in rule['condition'].items():
                condition &= (df[cond_col] == cond_val)

            special_count = condition.sum()
            if special_count > 0:
                # Apply mapping - แปลง Series เป็น numpy array เพื่อแก้ปัญหา "unhashable type: 'Series'"
                condition_mask = condition.values
                for map_col, map_val in rule['mapping'].items():
                    df.loc[condition_mask, map_col] = map_val
                self.log(f"    Applied to {special_count:,} รายการ")
        
        # บันทึกไฟล์
        output_file = os.path.join(self.paths["output"], self.config.OUTPUT_MAPPED_PRODUCT_FILE)
        df.to_csv(output_file, index=False, float_format="%.2f")
        
        self.log(f"บันทึกไฟล์: {output_file}")
        self.log(f"Mapping Product เสร็จสิ้น", df["REVENUE_VALUE"].sum())
        
        return df

    def _load_adj_data(self):
        """
        [ฟังก์ชันใหม่] โหลดไฟล์ ADJ รายเดือน และ ADJ YTD
        """
        self.log("--- Loading ADJ Data (Financial Income / Other Revenue) ---")
        
        # --- Load Monthly ADJ Files ---
        monthly_pattern = os.path.join(self.paths["input"], self.config.ADJ_MONTHLY_PATTERN)
        
        # [ START FIX ] ---
        # กรองไฟล์ YTD และ ADJ_GL_NT1 ออก
        self.log(f"ค้นหา pattern: {monthly_pattern}")
        monthly_files_all = glob.glob(monthly_pattern)
        monthly_files = []
        for f in monthly_files_all:
            filename_upper = os.path.basename(f).upper()
            if "YTD" not in filename_upper and "ADJ_GL_NT1" not in filename_upper:
                monthly_files.append(f)
            else:
                self.log(f"  ข้ามไฟล์: {os.path.basename(f)} (ไม่ตรงเงื่อนไข ADJ รายเดือน)")
        # [ END FIX ] ---
        
        df_adj_monthly = pd.DataFrame()
        if not monthly_files:
            self.log("⚠️  ไม่พบไฟล์ ADJ รายเดือน (TRN_REVENUE_ADJ_*.csv)")
        else:
            self.log(f"พบ {len(monthly_files)} ไฟล์ ADJ รายเดือน")
            for file in monthly_files:
                self.log(f"  กำลังอ่าน ADJ รายเดือน: {os.path.basename(file)}")
                try:
                    df = self.read_csv_robust(file)
                except Exception as e:
                     self.log(f"    ❌ อ่านไฟล์ {file} ไม่สำเร็จ: {e}")
                     continue
                
                df_adj_monthly = pd.concat([df_adj_monthly, df], ignore_index=True)
            
            if not df_adj_monthly.empty:
                # Clean monthly data
                df_adj_monthly = df_adj_monthly.dropna(subset=['REVENUE_VALUE'])
                df_adj_monthly["REVENUE_VALUE"] = pd.to_numeric(
                    df_adj_monthly["REVENUE_VALUE"], errors='coerce'
                ).fillna(0)

                # Filter ADJ data ให้เหลือเฉพาะเดือนที่กำหนดใน config
                if hasattr(self.config, 'end_month') and self.config.end_month:
                    if 'MONTH' in df_adj_monthly.columns:
                        original_count = len(df_adj_monthly)
                        df_adj_monthly['MONTH_INT'] = df_adj_monthly['MONTH'].astype(int)
                        df_adj_monthly = df_adj_monthly[
                            df_adj_monthly['MONTH_INT'] <= self.config.end_month
                        ]
                        df_adj_monthly = df_adj_monthly.drop(columns=['MONTH_INT'])
                        filtered_count = len(df_adj_monthly)
                        self.log(f"✓ Filter ADJ Monthly: {original_count} → {filtered_count} records (เดือน 1-{self.config.end_month})", "INFO")

                self.log(f"ยอดรวม ADJ รายเดือน (ที่โหลดได้): {df_adj_monthly['REVENUE_VALUE'].sum():,.2f}")

        # --- Load YTD ADJ File ---
        # เราต้องการไฟล์ YTD ล่าสุดเพียงไฟล์เดียว
        ytd_pattern = os.path.join(self.paths["input"], self.config.ADJ_YTD_PATTERN)
        ytd_files = glob.glob(ytd_pattern)
        
        df_adj_ytd = pd.DataFrame()
        if not ytd_files:
            self.log("⚠️  ไม่พบไฟล์ ADJ YTD (TRN_REVENUE_ADJ_YTD_*.csv)")
        else:
            latest_ytd_file = sorted(ytd_files)[-1] # เอาไฟล์ล่าสุดตามชื่อ
            self.log(f"พบ {len(ytd_files)} ไฟล์ YTD. ใช้ไฟล์ล่าสุด: {os.path.basename(latest_ytd_file)}")
            try:
                df_adj_ytd = self.read_csv_robust(latest_ytd_file)
            except Exception as e:
                 self.log(f"    ❌ อ่านไฟล์ YTD {latest_ytd_file} ไม่สำเร็จ: {e}")
            
            if not df_adj_ytd.empty:
                # Clean YTD data
                df_adj_ytd = df_adj_ytd.dropna(subset=['REVENUE_VALUE'])
                df_adj_ytd["REVENUE_VALUE"] = pd.to_numeric(
                    df_adj_ytd["REVENUE_VALUE"], errors='coerce'
                ).fillna(0)
                self.log(f"ยอดรวม ADJ YTD (ที่โหลดได้): {df_adj_ytd['REVENUE_VALUE'].sum():,.2f}")
            
        return df_adj_monthly, df_adj_ytd

    def _process_adj_data(self, df_adj):
        """
        [ฟังก์ชันใหม่] แปลง ADJ data ให้มีโครงสร้างเหมือน df_output
        """
        self.log("กำลังประมวลผล ADJ data...")
        if df_adj.empty:
            self.log("ไม่มีข้อมูล ADJ ให้ประมวลผล")
            return pd.DataFrame()
            
        df_adj['YEAR'] = df_adj['YEAR'].astype(str)
        df_adj['MONTH'] = df_adj['MONTH'].astype(int).astype(str).str.zfill(2)
        
        # [ START FIX ]
        # จัดการคอลัมน์ 'TYPE' และ 'REVENUE_TYPE' ที่อาจซ้ำซ้อนกัน
        if "REVENUE_TYPE" in df_adj.columns and "TYPE" in df_adj.columns:
            self.log("  พบทั้งคอลัมน์ 'TYPE' และ 'REVENUE_TYPE' ... กำลังรวมคอลัมน์")
            # เติมค่า 'TYPE' ที่ว่าง (NaN) ด้วยค่าจาก 'REVENUE_TYPE'
            df_adj['TYPE'] = df_adj['TYPE'].fillna(df_adj['REVENUE_TYPE'])
            # ลบคอลัมน์ 'REVENUE_TYPE' ที่ซ้ำซ้อนออก
            df_adj = df_adj.drop(columns=['REVENUE_TYPE'])
            self.log("  รวมคอลัมน์ TYPE/REVENUE_TYPE เรียบร้อย")
        elif "REVENUE_TYPE" in df_adj.columns:
            # ถ้ามีแค่ REVENUE_TYPE, ก็แค่เปลี่ยนชื่อ
            self.log("  พบ 'REVENUE_TYPE', ทำการเปลี่ยนชื่อเป็น 'TYPE'")
            df_adj = df_adj.rename(columns={"REVENUE_TYPE": "TYPE"})
        elif "TYPE" not in df_adj.columns:
            # ถ้าไม่มีทั้งคู่
            self.log("❌ ไฟล์ ADJ ขาดคอลัมน์ 'TYPE' หรือ 'REVENUE_TYPE'")
            return pd.DataFrame()
        # [ END FIX ]

        # กรองเฉพาะประเภทที่เราต้องการ (บรรทัดนี้คือจุดที่เคยเกิด Error)
        df_adj_filtered = df_adj[df_adj['TYPE'].isin([
            self.config.FINANCIAL_INCOME_NAME, 
            self.config.OTHER_REVENUE_ADJ_NAME
        ])].copy()
        
        if df_adj_filtered.empty:
            self.log("ไม่พบข้อมูล 'ผลตอบแทนทางการเงิน' หรือ 'รายได้อื่น' ในไฟล์ ADJ")
            return pd.DataFrame()

        # สร้าง DataFrame ใหม่ที่มีโครงสร้างเหมือน df_output
        df_processed = pd.DataFrame()
        df_processed['YEAR'] = df_adj_filtered['YEAR']
        df_processed['MONTH'] = df_adj_filtered['MONTH']
        df_processed['REVENUE_VALUE'] = df_adj_filtered['REVENUE_VALUE']
        
        # สร้างคอลัมน์ Hierarchy สำหรับใช้ในรายงาน
        df_processed['BUSINESS_GROUP'] = self.config.NEW_ADJ_BUSINESS_GROUP
        df_processed['SERVICE_GROUP'] = df_adj_filtered['TYPE'] # e.g., "ผลตอบแทนทางการเงิน"
        df_processed['PRODUCT_NAME'] = df_adj_filtered['TYPE']
        
        # สร้างคอลัมน์ Dummy/Placeholder เพื่อให้ concat กับ df_output ได้
        df_processed['ITEM'] = 'ADJ'
        df_processed['SUB_ITEM'] = 'ADJ'
        df_processed['PRODUCT_KEY'] = 'ADJ_' + df_adj_filtered['TYPE'].str.replace(' ', '_') # Unique key
        df_processed['SUB_PRODUCT_KEY'] = '1'
        df_processed['COST_CENTER'] = df_adj_filtered.get('COST_CENTER', 'NA')
        df_processed['CUSTOMER_GROUP_KEY'] = df_adj_filtered.get('CUSTOMER_GROUP_KEY', 'NA')
        df_processed['GL_CODE'] = 'ADJ' # Dummy GL
        df_processed['NT'] = 'NT'
        
        # สร้างคอลัมน์ Dummy จาก Master GL
        df_processed['REPORT_CODE'] = 'ADJ'
        df_processed['GL_NAME'] = 'ADJ'
        df_processed['GL_GROUP'] = 'ADJ'
        df_processed['หมวดบัญชี'] = 'ADJ' # สำคัญ: ต้องมีค่าเพื่อไม่ให้ Error ใน Step 4
        
        # สร้างคอลัมน์ Dummy จาก Master Product
        df_processed['SUB_PRODUCT_NAME'] = df_adj_filtered['TYPE']
        df_processed['REVENUE_GROUP_TYPE'] = 'ADJ'
        df_processed['BUSINESS'] = df_processed['ITEM'] + " " + df_processed['BUSINESS_GROUP']
        df_processed['SERVICE'] = df_processed['SUB_ITEM'] + " " + df_processed['SERVICE_GROUP']
        df_processed['PRODUCT'] = df_processed['PRODUCT_KEY'] + " " + df_processed['PRODUCT_NAME']
        df_processed['SUB_PRODUCT'] = df_processed['SUB_PRODUCT_KEY'] + " " + df_processed['SUB_PRODUCT_NAME']
        
        # สร้างคอลัมน์ Date และ Amount
        df_processed["dt"] = "01" + df_processed["MONTH"] + df_processed["YEAR"]
        df_processed["DATE"] = pd.to_datetime(df_processed["dt"], errors="coerce", format="%d%m%Y")
        
        df_processed['TYPE'] = 'รายได้'
        df_processed['AMOUNT'] = df_processed['REVENUE_VALUE']
        
        self.log(f"ประมวลผล ADJ data {len(df_processed)} แถวเรียบร้อย")
        return df_processed

    def step4_create_final_report(self, df):
        """ขั้นตอนที่ 4: สร้างรายงานสุดท้าย"""
        self.log("=" * 80)
        self.log("STEP 4: สร้างรายงานสุดท้าย")
        self.log("=" * 80)
        
        grand_total_before = df["REVENUE_VALUE"].sum()
        
        # อ่าน Master Product
        master_product_file = os.path.join(self.paths["master"], self.config.MASTER_PRODUCT_FILE)
        if not os.path.exists(master_product_file):
            raise FileNotFoundError(f"ไม่พบไฟล์ {master_product_file}")
        
        df_master_product = pd.read_csv(master_product_file, dtype=str)
        
        # สร้างคอลัมน์รวม
        df_master_product["BUSINESS"] = df_master_product["ITEM"] + " " + df_master_product["BUSINESS_GROUP"]
        df_master_product["SERVICE"] = df_master_product["SUB_ITEM"] + " " + df_master_product["SERVICE_GROUP"]
        df_master_product["PRODUCT"] = df_master_product["PRODUCT_KEY"] + " " + df_master_product["PRODUCT_NAME"]
        df_master_product["SUB_PRODUCT"] = df_master_product["SUB_PRODUCT_KEY"] + " " + df_master_product["SUB_PRODUCT_NAME"]
        
        df_master_product = df_master_product[[
            "PRODUCT_KEY", "PRODUCT_NAME", "SUB_PRODUCT_KEY", "SUB_PRODUCT_NAME",
            "ITEM", "BUSINESS_GROUP", "SUB_ITEM", "SERVICE_GROUP", 
            "REVENUE_GROUP_TYPE", "BUSINESS", "SERVICE", "PRODUCT", "SUB_PRODUCT"
        ]]
        df_master_product = df_master_product.drop_duplicates()
        
        self.log(f"Master Product: {len(df_master_product):,} รายการ")
        
        # อ่าน Master GL
        master_gl_file = os.path.join(self.paths["master"], self.config.MASTER_GL_FILE)
        if not os.path.exists(master_gl_file):
            raise FileNotFoundError(f"ไม่พบไฟล์ {master_gl_file}")
        
        df_master_gl = pd.read_csv(master_gl_file, converters={"GL_CODE_NT1": str})
        
        # เลือกเฉพาะ columns ที่ต้องการจาก Master GL
        df_master_gl = df_master_gl[[
            "GL_CODE_NT1", "GL_NAME_NT1", "REPORT_CODE", "GL_GROUP", "หมวดบัญชี"
        ]].copy()
        
        df_master_gl.rename(columns={"GL_NAME_NT1": "GL_NAME"}, inplace=True)
        
        self.log(f"Master GL: {len(df_master_gl):,} รายการ")
        
        # เพิ่มคอลัมน์เสริม
        df["NT"] = "NT"
        df["dt"] = "01" + df["MONTH"] + df["YEAR"]
        df["DATE"] = pd.to_datetime(df["dt"], errors="coerce", format="%d%m%Y")
        
        # Group by
        self.log("Group by ข้อมูล (NT1)...")
        df_output = df.groupby([
            "YEAR", "MONTH", "DATE", "COST_CENTER", "CUSTOMER_GROUP_KEY",
            "PRODUCT_KEY", "SUB_PRODUCT_KEY", "GL_CODE", "NT"
        ], dropna=False)["REVENUE_VALUE"].sum().reset_index()
        df_output = df_output.round(2)
        
        self.log(f"หลัง Group by (NT1): {len(df_output):,} แถว")
        
        # Merge กับ Master GL
        self.log("Merging (NT1) กับ Master GL...")
        
        gl_in_data = set(df_output["GL_CODE"].dropna().unique())
        gl_in_master = set(df_master_gl["GL_CODE_NT1"].dropna().unique())
        
        self.log(f"GL_CODE ในข้อมูล (NT1): {len(gl_in_data)} รหัส")
        self.log(f"GL_CODE_NT1 ใน Master: {len(gl_in_master)} รหัส")
        
        # หา GL ที่ไม่มีใน Master
        gl_not_in_master = gl_in_data - gl_in_master
        if len(gl_not_in_master) > 0:
            self.log(f"⚠️  พบ GL_CODE {len(gl_not_in_master)} รหัสที่ไม่มีใน Master:")
            for gl in sorted(list(gl_not_in_master))[:10]:
                count = (df_output["GL_CODE"] == gl).sum()
                amount = df_output[df_output["GL_CODE"] == gl]["REVENUE_VALUE"].sum()
                self.log(f"    - {gl}: {count:,} แถว, จำนวนเงิน {amount:,.2f}")
        
        df_output = pd.merge(
            df_output,
            df_master_gl,
            left_on="GL_CODE",
            right_on="GL_CODE_NT1",
            how="left"
        )
        
        df_output = df_output.drop(columns=["GL_CODE_NT1"])
        
        self.log(f"หลัง Merge GL (NT1): {len(df_output):,} แถว")
        
        # Group by again
        df_output = df_output.groupby([
            "YEAR", "MONTH", "DATE", "COST_CENTER", "CUSTOMER_GROUP_KEY",
            "PRODUCT_KEY", "SUB_PRODUCT_KEY", "REPORT_CODE", "GL_CODE",
            "GL_NAME", "GL_GROUP", "หมวดบัญชี", "NT"
        ], dropna=False)["REVENUE_VALUE"].sum().reset_index()
        df_output = df_output.round(2)
        
        df_output["TYPE"] = "รายได้"
        df_output["AMOUNT"] = df_output["REVENUE_VALUE"]
        
        # Merge กับ Master Product
        self.log("Merging (NT1) กับ Master Product...")

        # แปลง PRODUCT_KEY และ SUB_PRODUCT_KEY ให้เป็น string เพื่อ match กับ Master
        df_output["PRODUCT_KEY"] = df_output["PRODUCT_KEY"].astype(str)
        df_output["SUB_PRODUCT_KEY"] = df_output["SUB_PRODUCT_KEY"].astype(int).astype(str)

        df_output = pd.merge(
            df_output,
            df_master_product,
            on=["PRODUCT_KEY", "SUB_PRODUCT_KEY"],
            how="left"
        )
        
        self.log(f"หลัง Merge Product (NT1): {len(df_output):,} แถว")

        # --- [ START MODIFICATION ] ---
        # 1. กรอง "กลุ่มธุรกิจ รายได้อื่น" ออกจากข้อมูล NT1
        self.log(f"กำลังกรอง BUSINESS_GROUP: '{self.config.EXCLUDE_BUSINESS_GROUP}' ออกจากข้อมูล NT1")
        initial_count = len(df_output)
        
        if "BUSINESS_GROUP" not in df_output.columns:
            self.log("⚠️  ไม่พบคอลัมน์ BUSINESS_GROUP, ไม่สามารถกรองได้")
        else:
            df_output = df_output[df_output["BUSINESS_GROUP"] != self.config.EXCLUDE_BUSINESS_GROUP].copy()
            filtered_count = initial_count - len(df_output)
            self.log(f"  กรองออก {filtered_count:,} แถว")
            
        grand_total_after_filter = df_output["REVENUE_VALUE"].sum()
        self.log(f"ยอดรวม (NT1 data) หลังกรอง: {grand_total_after_filter:,.2f}")

        # 2. โหลดและประมวลผล ADJ data
        df_adj_monthly, df_adj_ytd = self._load_adj_data()
        
        # เก็บ YTD data ไว้ใช้ในส่วน __main__
        self.df_adj_ytd = df_adj_ytd 
        
        # 3. ประมวลผล ADJ data
        df_adj_processed = self._process_adj_data(df_adj_monthly)
        
        # 4. รวมข้อมูล NT1 ที่กรองแล้ว กับ ADJ data
        if df_adj_processed is not None and not df_adj_processed.empty:
            self.log("กำลังรวมข้อมูล NT1 และ ADJ...")
            df_output = pd.concat([df_output, df_adj_processed], ignore_index=True)
            self.log(f"จำนวนแถวทั้งหมดหลังรวม: {len(df_output):,}")
        else:
            self.log("ไม่พบข้อมูล ADJ รายเดือน, ใช้ข้อมูล NT1 ต่อไป")
        
        # --- [ END MODIFICATION ] ---

        
        # ตรวจสอบความถูกต้อง (ตอนนี้จะตรวจสอบข้อมูลที่รวม ADJ แล้ว)
        self.log("ตรวจสอบความถูกต้องของข้อมูล (NT1 + ADJ)...")
        
        # 'หมวดบัญชี' ของ ADJ จะเป็น 'ADJ' ซึ่งไม่ .isnull()
        missing_gl = df_output["หมวดบัญชี"].isnull() 
        # 'BUSINESS_GROUP' ของ ADJ จะมีค่า (NEW_ADJ_BUSINESS_GROUP) ซึ่งไม่ .isnull()
        missing_bu = df_output["BUSINESS_GROUP"].isnull()
        
        has_error = False
        
        if missing_gl.any():
            missing_count = missing_gl.sum()
            missing_amount = df_output[missing_gl]["REVENUE_VALUE"].sum()
            self.log(f"⚠️  พบข้อมูล GL (NT1) ที่ไม่มีใน Master: {missing_count:,} แถว (จำนวนเงิน: {missing_amount:,.2f})")
            
            problem_gl = df_output[missing_gl][["GL_CODE", "REVENUE_VALUE"]].groupby("GL_CODE").agg({
                "REVENUE_VALUE": "sum"
            }).reset_index().sort_values("REVENUE_VALUE", ascending=False)
            
            self.log(f"GL_CODE (NT1) ที่มีปัญหา (Top 10):")
            for idx, row in problem_gl.head(10).iterrows():
                self.log(f"  - {row['GL_CODE']}: {row['REVENUE_VALUE']:,.2f}")
            
            error_file = os.path.join(self.paths["final_output"], self.config.ERROR_GL_FILE)
            df_output[missing_gl].to_csv(error_file, index=False)
            self.log(f"บันทึก error log: {error_file}")
            has_error = True
        
        if missing_bu.any():
            missing_count = missing_bu.sum()
            missing_amount = df_output[missing_bu]["REVENUE_VALUE"].sum()
            self.log(f"⚠️  พบข้อมูล Product (NT1) ที่ไม่มีใน Master: {missing_count:,} แถว (จำนวนเงิน: {missing_amount:,.2f})")
            
            error_data = df_output.loc[missing_bu, [
                "MONTH", "PRODUCT_KEY", "SUB_PRODUCT_KEY", "REVENUE_VALUE"
            ]].drop_duplicates()
            
            self.log(f"Product (NT1) ที่มีปัญหา (Top 10):")
            for idx, row in error_data.head(10).iterrows():
                self.log(f"  - {row['PRODUCT_KEY']}-{row['SUB_PRODUCT_KEY']}: {row['REVENUE_VALUE']:,.2f}")
            
            error_file = os.path.join(self.paths["final_output"], self.config.ERROR_PRODUCT_FILE)
            error_data.to_csv(error_file, index=False)
            self.log(f"บันทึก error log: {error_file}")
            has_error = True
        
        if not has_error:
            self.log("✓ ข้อมูล GL และ Product (NT1) ถูกต้องครบถ้วน (ADJ ถูกข้ามการตรวจสอบนี้)")
        
        # ตรวจสอบยอดรวมและแจ้งเตือนถ้ามีรายการหายไป
        grand_total_after = df_output["REVENUE_VALUE"].sum()

        # คำนวณยอดที่หายไปจากการ mapping ไม่เจอ
        if has_error:
            error_file = os.path.join(self.paths["final_output"], self.config.ERROR_PRODUCT_FILE)
            if os.path.exists(error_file):
                df_error = pd.read_csv(error_file)
                if 'REVENUE_VALUE' in df_error.columns:
                    total_missing = df_error['REVENUE_VALUE'].sum()
                    missing_records = len(df_error)

                    self.log("=" * 80)
                    self.log("⚠️  MISSING DATA ALERT")
                    self.log("=" * 80)
                    self.log(f"❌ รายการที่ Mapping ไม่เจอ: {missing_records:,} รายการ")
                    self.log(f"❌ ยอดเงินที่หายไป: {total_missing:,.2f} บาท")
                    self.log(f"❌ เปอร์เซ็นต์ที่หายไป: {(total_missing/grand_total_before*100):.4f}%")
                    self.log("=" * 80)
                    self.log(f"📋 ตรวจสอบรายการที่ผิดพลาดได้ที่: {error_file}")
                    self.log("=" * 80)
                    self.log("🔧 แนะนำ: อัปเดต Master Product File ให้รองรับรายการใหม่")
                    self.log("=" * 80)

        self.log(f"Grand Total Before (NT1 Original): {grand_total_before:,.2f}")
        self.log(f"Grand Total After (NT1 Filtered + ADJ):  {grand_total_after:,.2f}")

        # บันทึกไฟล์สุดท้าย (ข้อมูลดิบที่รวม ADJ แล้ว)
        output_file = os.path.join(self.paths["final_output"], self.config.OUTPUT_FINAL_REPORT_FILE)
        df_output.to_csv(output_file, index=False)

        self.log(f"✓ บันทึกรายงานสุดท้าย (ข้อมูลดิบ): {output_file}")
        self.log(f"✓ ETL Pipeline เสร็จสมบูรณ์!", grand_total_after)

        return df_output
    
    def detect_historical_anomalies(self, df_final):
        """
        ตรวจสอบความผิดปกติย้อนหลังทุกเดือนแบบ Vectorization (เร็วกว่า Loop 100 เท่า)
        Returns: Dictionary {(type, identifier, date_string): status}
        """
        if not self.config.ENABLE_HISTORICAL_HIGHLIGHT:
            return {}

        self.log("=" * 80)
        self.log("HISTORICAL ANOMALY (VECTORIZED): กำลังตรวจสอบแบบรวดเร็ว...")
        
        heatmap_map = {} 
        
        # เตรียมข้อมูล: แปลงเดือนเป็น Int เพื่อการเรียงลำดับที่ถูกต้อง
        df_final = df_final.copy()
        df_final['MONTH_INT'] = df_final['MONTH'].astype(int)
        
        # วนลูปตามระดับ (Product, Service, Business, Grand Total)
        for level_name, level_config in self.config.ANOMALY_LEVELS.items():
            # [OPTION] ถ้าต้องการแค่ Product Level ให้ uncomment บรรทัดล่างนี้
            # if level_name != 'product': continue 
            
            group_by = level_config["group_by"]
            
            # 1. Prepare Wide Format DataFrame
            if group_by:
                df_grouped = df_final.groupby(group_by + ['MONTH_INT'], dropna=False)['REVENUE_VALUE'].sum().reset_index()
                df_pivot = df_grouped.pivot_table(index=group_by, columns='MONTH_INT', values='REVENUE_VALUE', fill_value=0)
            else:
                df_grouped = df_final.groupby(['MONTH_INT'], dropna=False)['REVENUE_VALUE'].sum().reset_index()
                df_pivot = df_grouped.pivot_table(columns='MONTH_INT', values='REVENUE_VALUE', fill_value=0)
                df_pivot.index = ["GRAND_TOTAL"]

            # 2. Vectorized Calculation
            df_calc = df_pivot.replace(0, np.nan)
            min_periods = self.config.ANOMALY_MIN_HISTORY
            
            # q1_matrix = df_calc.expanding(min_periods=min_periods, axis=1).quantile(0.25).shift(1, axis=1)
            # q3_matrix = df_calc.expanding(min_periods=min_periods, axis=1).quantile(0.75).shift(1, axis=1)
            
            # สลับแกน (T) -> คำนวณ (ไม่มี axis=1) -> สลับกลับ (T)
            q1_matrix = df_calc.T.expanding(min_periods=min_periods).quantile(0.25).shift(1).T
            q3_matrix = df_calc.T.expanding(min_periods=min_periods).quantile(0.75).shift(1).T
            iqr_matrix = q3_matrix - q1_matrix
            
            k = self.config.ANOMALY_IQR_MULTIPLIER
            upper_fence = q3_matrix + (k * iqr_matrix)
            lower_fence = q1_matrix - (k * iqr_matrix)
            lower_fence[lower_fence < 0] = 0 
            
            # 3. Generate Status Masks
            is_neg = df_pivot < 0
            is_high = (df_pivot > upper_fence) & (upper_fence.notna())
            is_low = (df_pivot < lower_fence) & (lower_fence.notna())
            
            # 4. Convert to Dictionary for Excel
            def add_to_map(mask_df, status_code):
                anomalies = mask_df.stack()
                anomalies = anomalies[anomalies] # Filter True only
                
                for idx, _ in anomalies.items():
                    # idx structure depend on group_by length. 
                    # For Product: (Biz, Svc, Key, Name, Month)
                    month_int = idx[-1]
                    row_keys = idx[:-1]
                    
                    # [FIX] สร้าง Key ให้ตรงกับ Excel Loop (ตัด Product Name ออก)
                    if level_name == 'product':
                        # row_keys = (Biz, Svc, ProdKey, ProdName) -> เอาแค่ 3 ตัวแรก
                        lookup_key = (row_keys[0], row_keys[1], row_keys[2]) 
                    elif level_name == 'service':
                        # row_keys = (Biz, Svc)
                        lookup_key = (row_keys[0], row_keys[1])
                    elif level_name == 'business':
                        lookup_key = row_keys[0]
                    else:
                        lookup_key = "GRAND_TOTAL"
                        
                    date_str = f"01/{month_int:02d}/{self.config.YEAR}"
                    heatmap_map[(level_name, lookup_key, date_str)] = status_code

            add_to_map(is_low, 'Low_Spike')
            add_to_map(is_high, 'High_Spike')
            add_to_map(is_neg, 'Negative_Value')
            
        self.log(f"  ประมวลผลเสร็จสิ้น พบจุดผิดปกติ {len(heatmap_map)} จุด")
        return heatmap_map
    
    def detect_anomalies(self, df_final):
        """
        ตรวจจับความผิดปกติในเดือนล่าสุด
        ทำการตรวจสอบใน 4 ระดับ: Product, Service, Business, Grand Total
        (ตอนนี้จะทำงานกับข้อมูลที่รวม ADJ แล้ว)
        """
        self.log("=" * 80)
        self.log("ANOMALY DETECTION: ตรวจจับความผิดปกติในเดือนล่าสุด")
        self.log("=" * 80)

        # หาเดือนล่าสุด - ใช้ end_month จาก config แทนการใช้ max จาก data
        df_final['MONTH_INT'] = df_final['MONTH'].astype(int)
        if hasattr(self.config, 'end_month') and self.config.end_month:
            latest_month = self.config.end_month
            self.log(f"ใช้เดือนจาก config: {latest_month} (end_month)", "INFO")
        else:
            latest_month = df_final['MONTH_INT'].max()
            self.log(f"ใช้เดือนจาก data: {latest_month} (max month)", "WARNING")

        self.log(f"เดือนล่าสุดที่ตรวจสอบ: {latest_month}")
        
        anomaly_results = {}
        
        for level_name, level_config in self.config.ANOMALY_LEVELS.items():
            self.log(f"\nตรวจสอบระดับ: {level_name.upper()}")
            
            group_by = level_config["group_by"]
            
            # Aggregate ข้อมูลตามระดับ
            if group_by:
                df_grouped = df_final.groupby(
                    group_by + ['MONTH_INT'], 
                    dropna=False
                )['REVENUE_VALUE'].sum().reset_index()
            else:
                # Grand total - รวมทั้งหมด
                df_grouped = df_final.groupby(
                    ['MONTH_INT'], 
                    dropna=False
                )['REVENUE_VALUE'].sum().reset_index()
                df_grouped['LEVEL'] = 'GRAND_TOTAL'
                group_by = ['LEVEL']
            
            # Pivot เป็น wide format (แถว = entity, คอลัมน์ = เดือน)
            if group_by:
                df_pivot = df_grouped.pivot_table(
                    index=group_by,
                    columns='MONTH_INT',
                    values='REVENUE_VALUE',
                    fill_value=0
                ).reset_index()
            else:
                # สำหรับ Grand Total
                df_pivot = df_grouped.pivot_table(
                    index=group_by,
                    columns='MONTH_INT',
                    values='REVENUE_VALUE',
                    fill_value=0
                ).reset_index()
            
            # สร้างชื่อคอลัมน์เดือน
            month_cols = [col for col in df_pivot.columns if isinstance(col, int)]
            month_cols.sort()
            
            if latest_month not in month_cols:
                self.log(f"⚠️  ไม่พบเดือน {latest_month} ในข้อมูล - ข้าม")
                continue
            
            # คอลัมน์เดือนล่าสุดและเดือนในอดีต
            latest_col = latest_month
            historical_cols = [m for m in month_cols if m < latest_month]
            
            if len(historical_cols) < self.config.ANOMALY_MIN_HISTORY:
                self.log(f"⚠️  ข้อมูลในอดีตไม่เพียงพอ ({len(historical_cols)} เดือน) - ข้าม")
                continue
            
            self.log(f"ตรวจสอบ {len(df_pivot)} รายการ")
            self.log(f"เดือนในอดีต: {historical_cols}")
            
            # ตรวจจับความผิดปกติ
            df_pivot['ANOMALY_STATUS'] = df_pivot.apply(
                lambda row: self._check_anomaly_row(
                    row, 
                    latest_col, 
                    historical_cols
                ), 
                axis=1
            )
            
            # เพิ่มข้อมูลเสริม
            df_pivot['LATEST_MONTH'] = latest_month
            df_pivot['LATEST_VALUE'] = df_pivot[latest_col]
            
            # คำนวณค่าเฉลี่ยในอดีต
            df_pivot['AVG_HISTORICAL'] = df_pivot[historical_cols].mean(axis=1)

            # [เพิ่มใหม่] 1. หาคอลัมน์สำหรับ Rolling Window (เช่น 3 เดือนล่าสุด)
            window = self.config.ANOMALY_ROLLING_WINDOW
            # ถ้าประวัติยาวกว่า window ให้ตัดเอาเฉพาะท้ายๆ, ถ้าสั้นกว่าก็เอาทั้งหมด
            rolling_cols = historical_cols[-window:] if len(historical_cols) >= window else historical_cols
            
            # [เพิ่มใหม่] 2. คำนวณ Rolling Average
            df_pivot['ROLLING_AVG'] = df_pivot[rolling_cols].mean(axis=1)
            
            # คำนวณ % change
            df_pivot['PCT_CHANGE'] = (
                (df_pivot['LATEST_VALUE'] - df_pivot['AVG_HISTORICAL']) / 
                df_pivot['AVG_HISTORICAL'].replace(0, np.nan) * 100
            ).fillna(0)

            # [เพิ่มใหม่] 3. คำนวณ % change (เทียบกับ Rolling Average)
            df_pivot['PCT_ROLLING'] = (
                (df_pivot['LATEST_VALUE'] - df_pivot['ROLLING_AVG']) / 
                df_pivot['ROLLING_AVG'].replace(0, np.nan) * 100
            ).fillna(0)
            
            # สรุปผล
            status_counts = df_pivot['ANOMALY_STATUS'].value_counts()
            self.log(f"ผลการตรวจสอบ:")
            for status, count in status_counts.items():
                self.log(f"  - {status}: {count} รายการ")

            # คำนวณ summary สำหรับ web interface
            summary = {
                'total_anomalies': int(status_counts.drop('Normal', errors='ignore').sum()),
                'high_spikes': int(status_counts.get('High_Spike', 0)),
                'low_dips': int(status_counts.get('Low_Dip', 0)),
                'new_items': int(status_counts.get('New_Item', 0)),
                'negative_values': int(status_counts.get('Negative_Value', 0)),
                'normal': int(status_counts.get('Normal', 0)),
                'total_records': len(df_pivot)
            }

            # เก็บผลลัพธ์พร้อม summary และ dataframe
            anomaly_results[level_name] = {
                'summary': summary,
                'dataframe': df_pivot,
                'status_counts': status_counts.to_dict()
            }

        return anomaly_results
    
    def _check_anomaly_row(self, row, latest_col, historical_cols):
        """
        ตรวจสอบความผิดปกติของแถวเดียว
        ใช้ IQR method (Tukey's Fences)
        """
        # ดึงค่าเดือนล่าสุด
        latest_val = row[latest_col]
        
        # 1. ตรวจสอบค่าติดลบ
        if latest_val < 0:
            return "Negative_Value"
        
        # ดึงค่าในอดีต
        history = row[historical_cols]
        
        # กรองเอาเฉพาะค่าปกติ (> 0)
        history_clean = history[history > 0]
        
        # 2. ตรวจสอบว่ามีข้อมูลพอหรือไม่
        if len(history_clean) < self.config.ANOMALY_MIN_HISTORY:
            # ถ้ามีค่าเดือนล่าสุด แต่ไม่มีประวัติ (หรือมีน้อยไป)
            if latest_val > 0:
                return "New_Item" # เราอาจจะเปลี่ยนเป็น "Not_Enough_Data"
            return "Not_Enough_Data"
        
        # 3. คำนวณ Robust Statistics (IQR)
        Q1 = history_clean.quantile(0.25)
        Q3 = history_clean.quantile(0.75)
        IQR = Q3 - Q1
        
        # 4. กรณีพิเศษ: ยอดในอดีตคงที่ (IQR = 0)
        if IQR == 0:
            # ถ้าค่าในอดีตเป็น 0 หมด และเดือนนี้มีค่า
            if Q1 == 0 and latest_val > 0:
                return "High_Spike" # ถือเป็น Spike
            return "Normal" if latest_val == Q1 else "Spike_vs_Constant"
        
        # 5. สร้างรั้ว (Tukey's Fences)
        k = self.config.ANOMALY_IQR_MULTIPLIER
        lower_fence = Q1 - (k * IQR)
        upper_fence = Q3 + (k * IQR)
        
        # 6. ตรวจสอบและส่งผล
        if latest_val > upper_fence:
            return "High_Spike"
        
        # ไม่ให้ค่าต่ำกว่า 0
        lower_fence_adjusted = max(0, lower_fence) 

        if latest_val < lower_fence_adjusted:
            return "Low_Spike"
        
        return "Normal"
    
    def create_anomaly_report_sheets(self, anomaly_results, output_file):
        """
        สร้าง sheet เพิ่มเติมใน Excel สำหรับ Anomaly Detection Report
        """
        self.log("=" * 80)
        self.log("สร้าง Anomaly Detection Report Sheets")
        self.log("=" * 80)
        
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.log("❌ ไม่พบ openpyxl, ไม่สามารถสร้าง Anomaly Sheets ได้")
            self.log("  โปรดติดตั้งด้วย: pip install openpyxl")
            return

        
        # เปิดไฟล์ Excel ที่สร้างไว้แล้ว
        try:
            wb = load_workbook(output_file)
        except Exception as e:
            self.log(f"❌ ไม่สามารถเปิดไฟล์ Excel ได้: {e}")
            return
        
        # สร้าง Summary Sheet
        self._create_anomaly_summary_sheet(wb, anomaly_results)

        # สร้าง Detail Sheets สำหรับแต่ละระดับ
        for level_name, level_data in anomaly_results.items():
            self._create_anomaly_detail_sheet(wb, level_name, level_data)
        
        # บันทึกไฟล์
        try:
            wb.save(output_file)
            self.log(f"✓ บันทึก Anomaly Report เรียบร้อย: {output_file}")
        except Exception as e:
            self.log(f"❌ ไม่สามารถบันทึกไฟล์ Excel ได้ (อาจจะเปิดค้างอยู่): {e}")
    
    def _create_anomaly_summary_sheet(self, wb, anomaly_results):
        """สร้าง Summary Sheet แสดงภาพรวมของการตรวจสอบ"""
        
        from openpyxl.styles import Font, PatternFill, Alignment

        if 'Anomaly Summary' in wb.sheetnames:
            try:
                del wb['Anomaly Summary']
            except Exception as e:
                self.log(f"  Warning: ไม่สามารถลบ sheet 'Anomaly Summary' เก่าได้: {e}")

        ws = wb.create_sheet('Anomaly Summary', 0) # สร้างไว้หน้าแรก
        
        # Header
        ws['A1'] = 'ANOMALY DETECTION SUMMARY'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        
        # Column headers
        headers = ['Level', 'Total Items', 'Normal', 'High Spike', 'Low Spike', 'Other Issues']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row = 4
        if not anomaly_results:
            ws.cell(row=row, column=1, value="ไม่มีผลลัพธ์การตรวจสอบ")
            return

        for level_name, level_data in anomaly_results.items():
            # เข้าถึง dataframe และ status_counts จาก dict structure ใหม่
            df_result = level_data.get('dataframe')
            if df_result is None:
                continue

            # ใช้ status_counts ที่คำนวณไว้แล้ว หรือคำนวณใหม่
            status_counts = level_data.get('status_counts', {})
            if not status_counts and 'ANOMALY_STATUS' in df_result.columns:
                status_counts = df_result['ANOMALY_STATUS'].value_counts().to_dict()

            ws.cell(row=row, column=1, value=level_name.upper())
            ws.cell(row=row, column=2, value=len(df_result))
            ws.cell(row=row, column=3, value=status_counts.get('Normal', 0))
            ws.cell(row=row, column=4, value=status_counts.get('High_Spike', 0))
            ws.cell(row=row, column=5, value=status_counts.get('Low_Spike', 0))

            other = (
                status_counts.get('Negative_Value', 0) +
                status_counts.get('Not_Enough_Data', 0) +
                status_counts.get('Spike_vs_Constant', 0) +
                status_counts.get('New_Item', 0)
            )
            ws.cell(row=row, column=6, value=other)
            
            # Highlight ถ้ามีปัญหา
            if (status_counts.get('High_Spike', 0) > 0 or 
                status_counts.get('Low_Spike', 0) > 0 or
                status_counts.get('Negative_Value', 0) > 0):
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'
                    )
            
            row += 1
        
        # ปรับความกว้างคอลัมน์
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
    
    def _create_anomaly_detail_sheet(self, wb, level_name, level_data):
        """สร้าง Detail Sheet สำหรับแต่ละระดับ"""

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # เข้าถึง dataframe จาก dict structure ใหม่
        df_result = level_data.get('dataframe')
        if df_result is None:
            return

        sheet_name = f'Anomaly_{level_name.title()}'

        if sheet_name in wb.sheetnames:
            try:
                del wb[sheet_name]
            except Exception as e:
                self.log(f"  Warning: ไม่สามารถลบ sheet '{sheet_name}' เก่าได้: {e}")

        ws = wb.create_sheet(sheet_name)

        # กรองเฉพาะรายการที่ผิดปกติ
        if 'ANOMALY_STATUS' in df_result.columns:
            df_anomaly = df_result[~df_result['ANOMALY_STATUS'].isin(['Normal', 'Not_Enough_Data'])].copy()
        else:
            df_anomaly = pd.DataFrame()  # Empty dataframe if no ANOMALY_STATUS column

        if len(df_anomaly) == 0:
            ws['A1'] = f'ไม่พบความผิดปกติในระดับ {level_name.upper()}'
            ws['A1'].font = Font(bold=True, size=12, color='008000')
            return
        
        # เรียงตาม absolute % change
        df_anomaly['ABS_PCT_CHANGE'] = df_anomaly['PCT_CHANGE'].abs()
        df_anomaly = df_anomaly.sort_values('ABS_PCT_CHANGE', ascending=False)
        
        # เลือกคอลัมน์ที่จะแสดง
        display_cols = []
        
        # เพิ่ม group columns
        group_cols = self.config.ANOMALY_LEVELS[level_name]["group_by"]
        if group_cols:
            display_cols.extend(group_cols)
        else:
            display_cols.append('LEVEL')
        
        display_cols.extend([
            'ANOMALY_STATUS', 'LATEST_VALUE', 'AVG_HISTORICAL', 'PCT_CHANGE', 'ROLLING_AVG', 'PCT_ROLLING'
        ])
        
        df_display = df_anomaly[display_cols].copy()
        
        # เขียนข้อมูลลง Excel
        # Headers
        for col_idx, col_name in enumerate(df_display.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        for row_idx, (_, row) in enumerate(df_display.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # จัดรูปแบบตัวเลข
                if col_idx > len(group_cols):
                    # เช็คชื่อคอลัมน์ว่ามีคำว่า PCT หรือไม่ เพื่อใส่ %
                    if 'PCT' in df_display.columns[col_idx-1]:
                        cell.number_format = '0.00"%"'
                    else:
                        cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                
                # Highlight ตาม status
                status = row['ANOMALY_STATUS']
                if status == 'High_Spike':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # Red
                elif status == 'Low_Spike':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') # Yellow
                elif status == 'Negative_Value':
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') # Bright Red
                    cell.font = Font(color='FFFFFF')
                elif status == 'Spike_vs_Constant':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # Red
                elif status == 'New_Item':
                    cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid') # Green
        
        # ปรับความกว้างคอลัมน์
        for col_idx, col_name in enumerate(df_display.columns, 1):
            width = 20
            if col_name in ('PRODUCT_NAME', 'SERVICE_GROUP', 'BUSINESS_GROUP'):
                width = 35
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    def run(self):
        """รัน ETL Pipeline ทั้งหมด พร้อม Anomaly Detection"""
        try:
            start_time = datetime.now()
            self.log(f"เริ่มต้น ETL Pipeline สำหรับปี {self.config.YEAR}")
            self.log(f"เวลาเริ่มต้น: {start_time}")
            
            # Step 1-4: ETL Process (Step 4 ถูกแก้ไขให้รวม ADJ data)
            df = self.step1_concat_revenue_files()

            # ============================================================
            # Step 0: Reconciliation (หลัง step1, ก่อน step2) <<< เพิ่มส่วนนี้
            # ============================================================
            try:
                self.step0_reconcile_revenue()
            except ReconciliationError as e:
                # แสดง error message แต่ไม่ print traceback
                self.log("\n❌ ETL Pipeline หยุดการทำงาน: Reconciliation Failed", "ERROR")
                self.log("💡 กรุณาตรวจสอบ log file สำหรับรายละเอียด", "INFO")
                # Return None แทน raise เพื่อลด traceback ซ้ำซ้อน
                return None, {}

            df = self.step2_mapping_cost_center(df)
            df = self.step3_mapping_product(df)
            df_final = self.step4_create_final_report(df)
            
            # Step 5: Anomaly Detection (ตอนนี้จะรันบนข้อมูล NT1 + ADJ)
            self.log("\n" + "=" * 80)
            self.log("STEP 5: Anomaly Detection")
            self.log("=" * 80)
            
            anomaly_results = self.detect_anomalies(df_final)
            
            # บันทึก anomaly results เป็น CSV
            for level_name, level_data in anomaly_results.items():
                # เข้าถึง dataframe จาก dict structure ใหม่
                df_result = level_data.get('dataframe')
                if df_result is None or df_result.empty:
                    self.log(f"ข้ามการบันทึก anomaly level '{level_name}' (ไม่มีข้อมูล)")
                    continue
                output_file = os.path.join(
                    self.paths["final_output"],
                    f"anomaly_{level_name}_{self.config.YEAR}.csv"
                )
                df_result.to_csv(output_file, index=False)
                self.log(f"บันทึกผล anomaly detection: {output_file}")
            
            end_time = datetime.now()
            duration = end_time - start_time
            
            self.log("=" * 80)
            self.log(f"✓ ETL Pipeline เสร็จสมบูรณ์!")
            self.log(f"เวลาที่ใช้: {duration}")
            self.log("=" * 80)
            
            return df_final, anomaly_results

        except Exception as e:
            self.log(f"❌ เกิดข้อผิดพลาดร้ายแรงใน ETL Pipeline: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def create_excel_report(self, df_result, anomaly_results, excel_output_file=None):
        """
        สร้าง Excel Report พร้อม Pivot Table และ Formatting

        Args:
            df_result: DataFrame ผลลัพธ์จาก ETL Pipeline
            anomaly_results: Dictionary ของ anomaly detection results
            excel_output_file: Optional path สำหรับไฟล์ Excel output (ถ้าไม่ระบุจะใช้ default)

        Returns:
            str: Path ของไฟล์ Excel ที่สร้างเสร็จ
        """
        # Import dependencies
        import numpy as np
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.logger.error("="*80)
            self.logger.error("❌ ไม่พบโมดูล openpyxl")
            self.logger.error("  โปรดติดตั้งด้วยคำสั่ง: pip install openpyxl")
            self.logger.error("  ไม่สามารถสร้างหรือจัดรูปแบบไฟล์ Excel ได้")
            self.logger.error("="*80)
            raise ImportError("openpyxl is required for creating Excel reports")

        # Detect historical anomalies
        historical_anomalies = {}
        if self.config.ENABLE_HISTORICAL_HIGHLIGHT:
            self.log("กำลังตรวจสอบ Historical Anomalies...")
            historical_anomalies = self.detect_historical_anomalies(df_result)

        # ดึงข้อมูล YTD ที่โหลดไว้ใน step 4
        df_adj_ytd = self.df_adj_ytd

        self.logger.info("\n" + "=" * 80)
        self.logger.info("สร้าง Excel Report...")
        self.logger.info("=" * 80)

        # กำหนดชื่อไฟล์ Excel
        if excel_output_file is None:
            # กำหนดเดือนที่ใช้ในชื่อไฟล์
            # 1. ใช้ end_month จาก config (ถ้ามี)
            # 2. ถ้าไม่มี ใช้เดือนที่มากที่สุดจากข้อมูล
            if hasattr(self.config, 'end_month') and self.config.end_month:
                report_month = self.config.end_month
                self.logger.info(f"ใช้เดือนจาก config: {report_month} (end_month)")
            else:
                # หาเดือนล่าสุดจากข้อมูล df_result
                if 'MONTH' in df_result.columns:
                    report_month = int(df_result['MONTH'].max())
                    self.logger.info(f"ใช้เดือนจากข้อมูล: {report_month} (max month)")
                else:
                    report_month = 12  # default ถ้าหาไม่เจอ
                    self.logger.warning(f"ไม่พบ MONTH column, ใช้ค่า default: {report_month}")

            # สร้างชื่อไฟล์แบบ revenue_report_YYYYMM.xlsx
            excel_output_file = os.path.join(
                self.paths["final_output"],
                f"revenue_report_{self.config.YEAR}{report_month:02d}.xlsx"
            )
        self.logger.info(f"ไฟล์ Excel ที่จะสร้าง: {excel_output_file}")

        # Aggregate ข้อมูลตาม hierarchy และเดือน
        df = df_result.copy()
        agg_data = df.groupby([
            'ITEM', 'BUSINESS_GROUP',
            'SUB_ITEM', 'SERVICE_GROUP',
            'PRODUCT_KEY', 'PRODUCT_NAME',
            'YEAR', 'MONTH'
        ])['AMOUNT'].sum().reset_index()

        # แปลง MONTH เป็น int
        agg_data['MONTH'] = agg_data['MONTH'].astype(int)

        # Filter ข้อมูลให้เหลือเฉพาะเดือนที่กำหนด (ป้องกัน Excel มี column เดือนที่ไม่ต้องการ)
        if hasattr(self.config, 'end_month') and self.config.end_month:
            original_count = len(agg_data)
            agg_data = agg_data[agg_data['MONTH'] <= self.config.end_month]
            filtered_count = len(agg_data)
            self.logger.info(f"  ✓ Filter Excel Report Data: {original_count} → {filtered_count} records (เดือน 1-{self.config.end_month})")

        # สร้างคอลัมน์วันที่ในรูปแบบ DD/MM/YYYY
        agg_data['DATE_STR'] = agg_data.apply(
            lambda row: f"01/{row['MONTH']:02d}/{row['YEAR']}",
            axis=1
        )

        # ========================================================================
        # Nested function: create_report
        # ========================================================================
        def create_report(agg_data, anomaly_results, df_adj_ytd, sort_ascending=True):
            """
            สร้างรายงานรายเดือน
            เวอร์ชันอัปเกรด:
            - รวม ANOMALY_STATUS
            - เพิ่มแถว "รวมรายได้จากการให้บริการ"
            - เพิ่มกลุ่ม "ผลตอบแทนทางการเงินและรายได้อื่น"
            - ใช้ YTD สำหรับคอลัมน์ "ผลรวม" ของกลุ่ม ADJ
            """

            self.logger.info(f"  กำลังสร้างรายงาน (sort_ascending={sort_ascending})...")

            # --- 1. สร้าง Anomaly Maps ---
            self.logger.info("    กำลังสร้าง Anomaly Maps...")
            prod_map, serv_map, biz_map, grand_total_status = {}, {}, {}, ''

            try:
                # Product Map
                if 'product' in anomaly_results and 'dataframe' in anomaly_results['product']:
                    df_prod = anomaly_results['product']['dataframe']
                    if 'ANOMALY_STATUS' in df_prod.columns:
                        prod_cols = ['BUSINESS_GROUP', 'SERVICE_GROUP', 'PRODUCT_KEY', 'ANOMALY_STATUS']
                        df_prod_subset = df_prod[prod_cols]
                        prod_map = df_prod_subset.set_index(['BUSINESS_GROUP', 'SERVICE_GROUP', 'PRODUCT_KEY'])['ANOMALY_STATUS'].to_dict()

                # Service Map
                if 'service' in anomaly_results and 'dataframe' in anomaly_results['service']:
                    df_serv = anomaly_results['service']['dataframe']
                    if 'ANOMALY_STATUS' in df_serv.columns:
                        serv_cols = ['BUSINESS_GROUP', 'SERVICE_GROUP', 'ANOMALY_STATUS']
                        df_serv_subset = df_serv[serv_cols]
                        serv_map = df_serv_subset.set_index(['BUSINESS_GROUP', 'SERVICE_GROUP'])['ANOMALY_STATUS'].to_dict()

                # Business Map
                if 'business' in anomaly_results and 'dataframe' in anomaly_results['business']:
                    df_biz = anomaly_results['business']['dataframe']
                    if 'ANOMALY_STATUS' in df_biz.columns:
                        biz_cols = ['BUSINESS_GROUP', 'ANOMALY_STATUS']
                        df_biz_subset = df_biz[biz_cols]
                        biz_map = df_biz_subset.set_index(['BUSINESS_GROUP'])['ANOMALY_STATUS'].to_dict()

                # Grand Total Status
                if 'grand_total' in anomaly_results and 'dataframe' in anomaly_results['grand_total']:
                    df_grand = anomaly_results['grand_total']['dataframe']
                    if 'ANOMALY_STATUS' in df_grand.columns and len(df_grand) > 0:
                        grand_total_status = df_grand['ANOMALY_STATUS'].values[0]

            except KeyError as e:
                self.logger.warning(f"    Warning: ไม่พบ anomaly key {e}, จะใช้ค่าว่าง")
                prod_map, serv_map, biz_map, grand_total_status = {}, {}, {}, ''
            except Exception as e:
                self.logger.error(f"    Error creating anomaly maps: {e}")
                prod_map, serv_map, biz_map, grand_total_status = {}, {}, {}, ''

            # --- 2. สร้าง YTD Map ---
            self.logger.info("    กำลังสร้าง YTD data map...")
            ytd_map = {}
            if df_adj_ytd is not None and not df_adj_ytd.empty:

                # จัดการคอลัมน์ 'TYPE' และ 'REVENUE_TYPE' ที่อาจซ้ำซ้อนกัน
                if "REVENUE_TYPE" in df_adj_ytd.columns and "TYPE" in df_adj_ytd.columns:
                    print("    YTD Map: พบทั้ง 'TYPE' และ 'REVENUE_TYPE', ทำการรวม...")
                    df_adj_ytd['TYPE'] = df_adj_ytd['TYPE'].fillna(df_adj_ytd['REVENUE_TYPE'])
                    df_adj_ytd = df_adj_ytd.drop(columns=['REVENUE_TYPE'])
                elif "REVENUE_TYPE" in df_adj_ytd.columns:
                    print("    YTD Map: พบ 'REVENUE_TYPE', ทำการเปลี่ยนชื่อเป็น 'TYPE'...")
                    df_adj_ytd = df_adj_ytd.rename(columns={"REVENUE_TYPE": "TYPE"})

                if "TYPE" in df_adj_ytd.columns:
                    # Sum by TYPE เผื่อมีหลายแถว
                    ytd_summary = df_adj_ytd.groupby("TYPE")["REVENUE_VALUE"].sum()
                    ytd_map[Config.FINANCIAL_INCOME_NAME] = ytd_summary.get(Config.FINANCIAL_INCOME_NAME, 0)
                    ytd_map[Config.OTHER_REVENUE_ADJ_NAME] = ytd_summary.get(Config.OTHER_REVENUE_ADJ_NAME, 0)
                else:
                    self.logger.warning("    Warning: ไม่พบคอลัมน์ 'TYPE' ในไฟล์ YTD")

            print(f"    YTD Map: {ytd_map}")

            # --- 3. เรียงข้อมูลและ Pivot ---
            if sort_ascending:
                agg_data_sorted = agg_data.sort_values(['YEAR', 'MONTH'], ascending=[True, True])
            else:
                agg_data_sorted = agg_data.sort_values(['YEAR', 'MONTH'], ascending=[True, False])

            pivot = agg_data_sorted.pivot_table(
                index=['ITEM', 'BUSINESS_GROUP', 'SUB_ITEM', 'SERVICE_GROUP', 'PRODUCT_KEY', 'PRODUCT_NAME'],
                columns='DATE_STR',
                values='AMOUNT',
                aggfunc='sum',
                fill_value=0
            )

            month_cols = agg_data_sorted['DATE_STR'].unique().tolist()
            pivot = pivot[month_cols]

            # คำนวณผลรวม (Sum(เดือน)) สำหรับทุกแถว
            pivot.insert(0, 'ผลรวม', pivot.sum(axis=1))
            result = pivot.reset_index()

            # --- 4. สร้าง DataFrame สุดท้ายพร้อม Status ---
            rows = []

            # แยกกลุ่ม NT1 ปกติ และกลุ่ม ADJ
            standard_groups_df = result[result['BUSINESS_GROUP'] != Config.NEW_ADJ_BUSINESS_GROUP]
            adj_group_df = result[result['BUSINESS_GROUP'] == Config.NEW_ADJ_BUSINESS_GROUP]

            grouped_items = standard_groups_df.groupby(['ITEM', 'BUSINESS_GROUP'])

            # ตัวแปรสำหรับเก็บข้อมูลเพื่อคำนวณ "รวมรายได้จากการให้บริการ"
            total_service_revenue_data = pd.DataFrame()

            # --- 4a. ประมวลผลกลุ่มธุรกิจปกติ (NT1) ---
            print("    ประมวลผลกลุ่มธุรกิจปกติ (NT1)")
            for (item, business_group), item_data in grouped_items:
                grouped_sub_items = item_data.groupby(['SUB_ITEM', 'SERVICE_GROUP'])

                for (sub_item, service_group), sub_item_data in grouped_sub_items:
                    # เพิ่มแถวของแต่ละ product
                    for _, row in sub_item_data.iterrows():

                        prod_key = (row['BUSINESS_GROUP'], row['SERVICE_GROUP'], row['PRODUCT_KEY'])
                        prod_status = prod_map.get(prod_key, '')

                        row_dict = {
                            'กลุ่มธุรกิจ': row['BUSINESS_GROUP'],
                            'กลุ่มบริการ': row['SERVICE_GROUP'],
                            'รหัสบริการ': row['PRODUCT_KEY'],
                            'ชื่อบริการ': row['PRODUCT_NAME'],
                        }

                        if sort_ascending:
                            row_dict['ผลรวม'] = row['ผลรวม']
                            for col in month_cols:
                                row_dict[col] = row[col]
                            row_dict['ANOMALY_STATUS'] = prod_status
                        else:
                            row_dict['ผลรวม'] = row['ผลรวม']
                            row_dict['ANOMALY_STATUS'] = prod_status
                            for col in month_cols:
                                row_dict[col] = row[col]

                        rows.append(row_dict)

                    # เพิ่มแถวผลรวมของกลุ่มบริการ
                    serv_key = (business_group, service_group)
                    serv_status = serv_map.get(serv_key, '')

                    sum_row = {
                        'กลุ่มธุรกิจ': '',
                        'กลุ่มบริการ': f'รวม {service_group}',
                        'รหัสบริการ': '',
                        'ชื่อบริการ': '',
                    }

                    if sort_ascending:
                        sum_row['ผลรวม'] = sub_item_data['ผลรวม'].sum()
                        for col in month_cols:
                            sum_row[col] = sub_item_data[col].sum()
                        sum_row['ANOMALY_STATUS'] = serv_status
                    else:
                        sum_row['ผลรวม'] = sub_item_data['ผลรวม'].sum()
                        sum_row['ANOMALY_STATUS'] = serv_status
                        for col in month_cols:
                            sum_row[col] = sub_item_data[col].sum()

                    rows.append(sum_row)

                    # เก็บข้อมูลสำหรับ "รวมรายได้จากการให้บริการ"
                    if service_group != Config.NON_TELECOM_SERVICE_GROUP:
                        total_service_revenue_data = pd.concat([
                            total_service_revenue_data,
                            sub_item_data
                        ])

                # เพิ่มแถวผลรวมของกลุ่มธุรกิจ
                biz_key = business_group
                biz_status = biz_map.get(biz_key, '')

                sum_row = {
                    'กลุ่มธุรกิจ': f'รวม {business_group}',
                    'กลุ่มบริการ': '',
                    'รหัสบริการ': '',
                    'ชื่อบริการ': '',
                }

                if sort_ascending:
                    sum_row['ผลรวม'] = item_data['ผลรวม'].sum()
                    for col in month_cols:
                        sum_row[col] = item_data[col].sum()
                    sum_row['ANOMALY_STATUS'] = biz_status
                else:
                    sum_row['ผลรวม'] = item_data['ผลรวม'].sum()
                    sum_row['ANOMALY_STATUS'] = biz_status
                    for col in month_cols:
                        sum_row[col] = item_data[col].sum()

                rows.append(sum_row)

            # --- 4b. เพิ่มแถว "รวมรายได้จากการให้บริการ" ---
            print("    เพิ่มแถว 'รวมรายได้จากการให้บริการ'")
            sum_row_service = {
                'กลุ่มธุรกิจ': 'รวมรายได้จากการให้บริการ',
                'กลุ่มบริการ': '',
                'รหัสบริการ': '',
                'ชื่อบริการ': '',
            }

            if sort_ascending:
                sum_row_service['ผลรวม'] = total_service_revenue_data['ผลรวม'].sum()
                for col in month_cols:
                    sum_row_service[col] = total_service_revenue_data[col].sum()
                sum_row_service['ANOMALY_STATUS'] = ''
            else:
                sum_row_service['ผลรวม'] = total_service_revenue_data['ผลรวม'].sum()
                sum_row_service['ANOMALY_STATUS'] = ''
                for col in month_cols:
                    sum_row_service[col] = total_service_revenue_data[col].sum()

            rows.append(sum_row_service)

            # --- 4c. ประมวลผลกลุ่ม ADJ ---
            print("    ประมวลผลกลุ่ม 'ผลตอบแทนทางการเงินและรายได้อื่น'")
            if not adj_group_df.empty:
                # เพิ่มแถวหัวข้อกลุ่ม
                adj_header_row = {
                    'กลุ่มธุรกิจ': Config.NEW_ADJ_BUSINESS_GROUP,
                    'กลุ่มบริการ': '', 'รหัสบริการ': '', 'ชื่อบริการ': ''
                }
                if sort_ascending:
                    adj_header_row['ผลรวม'] = ''
                    for col in month_cols: adj_header_row[col] = ''
                    adj_header_row['ANOMALY_STATUS'] = ''
                else:
                    adj_header_row['ผลรวม'] = ''
                    adj_header_row['ANOMALY_STATUS'] = ''
                    for col in month_cols: adj_header_row[col] = ''
                rows.append(adj_header_row)

                # เรียงลำดับ
                adj_group_df = adj_group_df.sort_values(by='SERVICE_GROUP', ascending=True)

                for _, row in adj_group_df.iterrows():
                    service_group = row['SERVICE_GROUP']

                    serv_key = (Config.NEW_ADJ_BUSINESS_GROUP, service_group)
                    serv_status = serv_map.get(serv_key, '')

                    row_dict = {
                        'กลุ่มธุรกิจ': '',
                        'กลุ่มบริการ': service_group,
                        'รหัสบริการ': '',
                        'ชื่อบริการ': '',
                    }

                    # ใช้ค่า YTD สำหรับคอลัมน์ "ผลรวม"
                    ytd_value = ytd_map.get(service_group, 0)

                    if sort_ascending:
                        row_dict['ผลรวม'] = ytd_value
                        for col in month_cols:
                            row_dict[col] = row[col]
                        row_dict['ANOMALY_STATUS'] = serv_status
                    else:
                        row_dict['ผลรวม'] = ytd_value
                        row_dict['ANOMALY_STATUS'] = serv_status
                        for col in month_cols:
                            row_dict[col] = row[col]

                    rows.append(row_dict)

                # เพิ่มแถว "รวมผลตอบแทนทางการเงินและรายได้อื่น"
                sum_adj_total_ytd = sum(ytd_map.values())

                sum_row_adj = {
                    'กลุ่มธุรกิจ': '',
                    'กลุ่มบริการ': 'รวมผลตอบแทนทางการเงินและรายได้อื่น',
                    'รหัสบริการ': '',
                    'ชื่อบริการ': '',
                }

                biz_key = Config.NEW_ADJ_BUSINESS_GROUP
                biz_status = biz_map.get(biz_key, '')

                if sort_ascending:
                    sum_row_adj['ผลรวม'] = sum_adj_total_ytd
                    for col in month_cols:
                        sum_row_adj[col] = adj_group_df[col].sum()
                    sum_row_adj['ANOMALY_STATUS'] = biz_status
                else:
                    sum_row_adj['ผลรวม'] = sum_adj_total_ytd
                    sum_row_adj['ANOMALY_STATUS'] = biz_status
                    for col in month_cols:
                        sum_row_adj[col] = adj_group_df[col].sum()

                rows.append(sum_row_adj)

            # --- 4d. เพิ่มแถวรวมทั้งสิ้น ---
            print("    เพิ่มแถว 'รวมทั้งสิ้น'")

            nt1_total = standard_groups_df['ผลรวม'].sum()
            adj_total_ytd = sum(ytd_map.values())
            report_grand_total = nt1_total + adj_total_ytd

            sum_row = {
                'กลุ่มธุรกิจ': 'รวมทั้งสิ้น',
                'กลุ่มบริการ': '',
                'รหัสบริการ': '',
                'ชื่อบริการ': '',
            }

            if sort_ascending:
                sum_row['ผลรวม'] = report_grand_total
                for col in month_cols:
                    sum_row[col] = result[col].sum()
                sum_row['ANOMALY_STATUS'] = grand_total_status
            else:
                sum_row['ผลรวม'] = report_grand_total
                sum_row['ANOMALY_STATUS'] = grand_total_status
                for col in month_cols:
                    sum_row[col] = result[col].sum()

            rows.append(sum_row)

            return pd.DataFrame(rows)

        # ========================================================================
        # Nested function: format_excel
        # ========================================================================
        def format_excel(filename, anomaly_map=None):
            """จัดรูปแบบ Excel ให้สวยงาม"""
            self.logger.info("กำลังจัดรูปแบบ Excel...")

            # โหลดค่า font จาก config
            FONT_NAME = self.config.EXCEL_FONT_NAME
            FONT_SIZE = self.config.EXCEL_FONT_SIZE

            try:
                wb = load_workbook(filename)
            except Exception as e:
                self.logger.error(f"  ❌ ไม่สามารถเปิดไฟล์ Excel เพื่อจัดรูปแบบได้: {e}")
                return

            for sheet_name in wb.sheetnames:
                # ข้าม anomaly sheets
                if sheet_name.startswith('Anomaly') or sheet_name == 'Anomaly Summary':
                    continue

                ws = wb[sheet_name]

                # จัดรูปแบบ header
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=FONT_SIZE)

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # จัดรูปแบบตัวเลข
                text_columns = {'กลุ่มธุรกิจ', 'กลุ่มบริการ', 'รหัสบริการ', 'ชื่อบริการ', 'ANOMALY_STATUS'}

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
                    for col_idx, cell in enumerate(row, 1):
                        header_value = ws.cell(row=1, column=col_idx).value

                        if header_value not in text_columns:
                            if cell.value and isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')

                # กำหนด default font สำหรับทุก cell ที่ยังไม่ได้กำหนด
                default_font = Font(name=FONT_NAME, size=FONT_SIZE)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        if cell.font.name == 'Calibri':  # default font ของ Excel
                            cell.font = default_font

                # กำหนดสีสำหรับแต่ละระดับ
                service_group_fill = PatternFill(start_color='E8F1F8', end_color='E8F1F8', fill_type='solid')
                service_group_font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE)

                business_group_fill = PatternFill(start_color='D0E2F0', end_color='D0E2F0', fill_type='solid')
                business_group_font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE)

                grand_total_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
                grand_total_font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE)

                total_service_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
                total_service_font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE, color='006100')

                adj_group_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                adj_group_font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE)

                adj_item_fill = PatternFill(start_color='FDF7F4', end_color='FDF7F4', fill_type='solid')
                adj_item_font = Font(name=FONT_NAME, bold=False, size=FONT_SIZE)

                adj_total_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
                adj_total_font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE)

                # วนลูป highlight แถวตามประเภท
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    business_value = row[0].value
                    service_value = row[1].value

                    if business_value and isinstance(business_value, str):
                        if business_value.startswith('รวมทั้งสิ้น'):
                            for cell in row:
                                cell.fill = grand_total_fill
                                cell.font = grand_total_font
                        elif business_value.startswith('รวมรายได้จากการให้บริการ'):
                            for cell in row:
                                cell.fill = total_service_fill
                                cell.font = total_service_font
                        elif business_value == Config.NEW_ADJ_BUSINESS_GROUP:
                            for cell in row:
                                cell.fill = adj_group_fill
                                cell.font = adj_group_font
                        elif business_value.startswith('รวม '):
                            for cell in row:
                                cell.fill = business_group_fill
                                cell.font = business_group_font

                    elif service_value and isinstance(service_value, str):
                        if service_value.startswith('รวม '):
                            for cell in row:
                                cell.fill = service_group_fill
                                cell.font = service_group_font
                        elif service_value == 'รวมผลตอบแทนทางการเงินและรายได้อื่น':
                            for cell in row:
                                cell.fill = adj_total_fill
                                cell.font = adj_total_font
                        elif service_value in (Config.FINANCIAL_INCOME_NAME, Config.OTHER_REVENUE_ADJ_NAME):
                             for cell in row:
                                cell.fill = adj_item_fill
                                cell.font = adj_item_font

                # ปรับความกว้างของคอลัมน์
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 35
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 40

                for col_idx in range(5, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15

                ws.freeze_panes = 'E2'

                # Format ANOMALY_STATUS column
                anomaly_col_idx = None
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == 'ANOMALY_STATUS':
                        anomaly_col_idx = idx
                        break

                if anomaly_col_idx:
                    try:
                        anomaly_col_letter = get_column_letter(anomaly_col_idx)
                        ws.column_dimensions[anomaly_col_letter].width = 20

                        header_cell = ws.cell(row=1, column=anomaly_col_idx)
                        header_cell.alignment = Alignment(horizontal='center', vertical='center')

                        spike_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        dip_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                        neg_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=anomaly_col_idx, max_col=anomaly_col_idx):
                            cell = row[0]
                            cell.alignment = Alignment(horizontal='center')

                            if cell.value == 'High_Spike':
                                cell.fill = spike_fill
                            elif cell.value == 'Low_Spike':
                                cell.fill = dip_fill
                            elif cell.value == 'Negative_Value':
                                cell.fill = neg_fill
                            elif cell.value == 'Spike_vs_Constant':
                                cell.fill = spike_fill
                            elif cell.value == 'New_Item':
                                cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')

                    except Exception as e:
                        self.logger.warning(f"Warning: ไม่สามารถจัดรูปแบบคอลัมน์ Anomaly ได้: {e}")

                # Highlight Cell Anomaly
                if anomaly_map:
                    print(f"  Applying anomaly highlights to {sheet_name}...")

                    col_date_map = {}
                    for cell in ws[1]:
                        if cell.value and '/' in str(cell.value):
                            col_date_map[cell.column] = str(cell.value)

                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        b_group = row[0].value
                        s_group = row[1].value
                        p_key = row[2].value

                        row_type = None
                        lookup_key = None

                        if p_key:
                             row_type = 'product'
                             lookup_key = (b_group, s_group, p_key)

                        if row_type == 'product' and lookup_key:
                            for col_idx, date_str in col_date_map.items():
                                map_key = (row_type, lookup_key, date_str)

                                if map_key in anomaly_map:
                                    status = anomaly_map[map_key]
                                    cell_to_color = ws.cell(row=row[0].row, column=col_idx)

                                    if status == 'High_Spike':
                                        cell_to_color.fill = spike_fill
                                    elif status == 'Low_Spike':
                                        cell_to_color.fill = dip_fill
                                    elif status == 'Negative_Value':
                                        cell_to_color.fill = neg_fill
                                        cell_to_color.font = Font(color='FFFFFF')

            try:
                wb.save(filename)
                self.logger.success(f"จัดรูปแบบ Excel เรียบร้อยแล้ว")
            except Exception as e:
                self.logger.error(f"❌ ไม่สามารถบันทึกไฟล์ Excel ที่จัดรูปแบบแล้วได้ (อาจจะเปิดค้างอยู่): {e}")

        # ========================================================================
        # สร้างรายงาน 2 แบบ
        # ========================================================================
        self.logger.info("กำลังสร้างรายงาน...")
        report_asc = create_report(agg_data, anomaly_results, df_adj_ytd, sort_ascending=True)
        report_desc = create_report(agg_data, anomaly_results, df_adj_ytd, sort_ascending=False)

        # บันทึกเป็น Excel
        try:
            with pd.ExcelWriter(excel_output_file, engine='openpyxl') as writer:
                report_asc.to_excel(writer, sheet_name='เรียงเดือนน้อย-มาก', index=False)
                report_desc.to_excel(writer, sheet_name='เรียงเดือนมาก-น้อย', index=False)
            self.logger.success(f"สร้างรายงานเรียบร้อยแล้ว: {excel_output_file}")
        except Exception as e:
            self.logger.error(f"❌ ไม่สามารถบันทึกไฟล์ Excel ได้ (อาจจะเปิดค้างอยู่): {e}")
            raise

        # NOTE: Reconciliation Sheet ถูกลบออกตามคำขอของผู้ใช้
        # ระบบจะแจ้งเตือน mapping errors ผ่าน console แทน

        # Format Excel
        format_excel(excel_output_file, anomaly_map=historical_anomalies)

        # เพิ่ม Anomaly Detection Sheets
        print("\n" + "=" * 80)
        self.logger.info("เพิ่ม Anomaly Detection Report...")
        print("=" * 80)

        try:
            self.create_anomaly_report_sheets(anomaly_results, excel_output_file)
        except Exception as e:
            self.logger.error(f"❌ เกิดข้อผิดพลาดในการเพิ่ม Anomaly Sheets: {str(e)}")
            import traceback
            traceback.print_exc()

        # สรุปผลลัพธ์
        print("\n" + "=" * 80)
        self.logger.info("สรุปข้อมูล:")
        self.logger.info("=" * 80)
        self.logger.info(f"จำนวนแถวทั้งหมด (NT1 + ADJ): {len(df_result):,}")
        self.logger.info(f"ยอดรวมทั้งหมด (NT1 + ADJ): {df_result['REVENUE_VALUE'].sum():,.2f}")
        self.logger.info(f"จำนวนเดือน: {df_result['MONTH'].nunique()}")
        print(f"จำนวน Product (รวม ADJ): {df_result['PRODUCT_KEY'].nunique()}")

        print("\n" + "=" * 80)
        print("Anomaly Detection Summary:")
        self.logger.info("=" * 80)
        if not anomaly_results:
            print("  ไม่มีผลลัพธ์การตรวจสอบ Anomaly (อาจมีข้อมูลไม่พอ)")
        else:
            for level_name, level_data in anomaly_results.items():
                # เข้าถึง dataframe และ status_counts จาก dict structure
                df_anomaly = level_data.get('dataframe')
                if df_anomaly is None:
                    continue

                # ใช้ status_counts ที่คำนวณไว้แล้ว
                status_counts = level_data.get('status_counts', {})
                if not status_counts and 'ANOMALY_STATUS' in df_anomaly.columns:
                    status_counts = df_anomaly['ANOMALY_STATUS'].value_counts().to_dict()

                total = len(df_anomaly)
                normal = status_counts.get('Normal', 0) + status_counts.get('Not_Enough_Data', 0)
                anomalies = total - normal
                print(f"{level_name.upper():15s}: {anomalies:4d} / {total:4d} anomalies detected")

        print("\n✓ เสร็จสมบูรณ์!")
        print(f"Excel Report: {excel_output_file}")
        print(f"  - Sheet 1: เรียงเดือนน้อย-มาก")
        print(f"  - Sheet 2: เรียงเดือนมาก-น้อย")

        return excel_output_file


# ============================================================================
# Main Execution
# ============================================================================
if __name__ == "__main__":
    from datetime import datetime

    print("="*80)
    print("Revenue ETL System - Standalone Execution")
    print("="*80)

    # สร้าง ETL instance และรัน Pipeline
    etl = RevenueETL()
    df_result, anomaly_results = etl.run()

    # สร้าง Excel Report
    excel_file = etl.create_excel_report(df_result, anomaly_results)
